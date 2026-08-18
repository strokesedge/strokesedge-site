#!/usr/bin/env python3
"""
weekly_model_pipeline.py — StrokesEdge unattended weekly model pipeline.

Runs every 3 hours, Monday through Friday, via Windows Task Scheduler (see
SETUP INSTRUCTIONS at the bottom of this file). Each firing detects every
PGA Tour event happening "this week" (the main event plus any concurrent
opposite-field event), and advances each one independently through:

    field pull -> weight proposal (Claude API) -> human approval (file+email)
    -> L1/L2 regression -> odds-readiness check -> L3 value screen
    -> Excel workbook -> email the finished workbook as an attachment

Nobody triggers this by typing into a Claude Code session — it is a state
machine that checks a per-tournament state.json on every firing and does
whatever the next unblocked step is, then exits. See
weekly-model/CLAUDE.md for the full design rationale; this docstring only
covers what a reader of the code itself needs.

═══════════════════════════════════════════════════════════════════════════
DESIGN DECISIONS CARRIED OVER FROM weekly_course_update.py — READ FIRST
═══════════════════════════════════════════════════════════════════════════
This script reuses that script's proven patterns rather than reinventing
them: the lock-file concurrency guard, the logging setup (UTF-8 file
handler + degraded console output for Task Scheduler's non-UTF-8 codepage),
and — most importantly — the "hold locally, email a summary, human does
the final call" philosophy. The weight-approval step never auto-approves on
a timeout; if Brian is unreachable a whole tournament week, the accepted
outcome is no workbook that week, not an unreviewed model going out. See
the AUTO_PUSH discussion in weekly_course_update.py's docstring for the
same philosophy applied to that script's git-push step.

DATA GOLF, NOT PGA TOUR, IS THE PRIMARY DATA SOURCE:
  Verified this session that Data Golf's "True SG" and PGA Tour's official
  SG are NOT the same stat, even at similar scale (Scheffler: PGA Tour
  SG:APR +0.545 season vs Data Golf sg_app +1.028 — nearly 2x). This
  pipeline runs entirely on Data Golf's own scale. A PGA Tour CSV
  supplement is optional (checked for automatically, used only for stats
  Data Golf can't match at the same granularity — see PGA_SUPPLEMENT_FILE
  below) and is NEVER blended into the same column as a Data-Golf-sourced
  stat.

CONCURRENT EVENTS (e.g. The Open Championship + Corales Puntacana
Championship, same week — a real, live example seen while building this,
not a hypothetical):
  Data Golf's `tour` parameter distinguishes `pga` (the main/featured event
  that week) from `opp` (opposite-field companion event, when one exists).
  Verified directly: field-updates/betting-tools with tour=opp correctly
  isolate the alternate event's own field/odds, separate from tour=pga's.
  get-schedule itself does NOT accept tour=opp (tested, returns an error)
  — both events appear together in the tour=pga schedule listing; it's
  specifically field-updates/betting-tools where tour=opp pulls the
  alternate event's own data. See detect_events_this_week().

WEIGHT-SETTING IS A HELD-FOR-REVIEW LLM JUDGMENT CALL, LIKE THE COURSE
ANALYSIS TEXT IN weekly_course_update.py:
  propose_weights() calls the Claude API with course facts assembled from
  Wikipedia (same fallback source and same "mark TBD, don't invent"
  discipline as weekly_course_update.py's wikipedia_par_yardage()) and a
  fixed whitelist of Data-Golf-computable factors. Like that script's course
  analysis, this call has no live web-search grounding — it is a proposal,
  written to a file and emailed for a human sanity-check, never used
  unreviewed. Uses claude-sonnet-5, not haiku — a mispriced L2 gate
  threshold is a more consequential mistake than imprecise course-profile
  prose, so this call is deliberately not the cheap/fast model.

WHAT THIS SCRIPT DOES NOT DO:
  Does not touch strokesedge-site. Does not talk to "the workbook sender"
  (a separate Google Apps Script in Google's cloud, out of scope here).
  Its job ends at emailing Brian the finished .xlsx as an attachment.

OPEN QUESTIONS NOT YET VALIDATED AGAINST A REAL WEEK (see weekly-model/
CLAUDE.md "Open Questions" for the full list) — flagged in code comments
at the relevant spot rather than repeated here:
  - Whether a failed/empty tour=opp response reliably means "no concurrent
    event" vs. a transient API issue.
  - Whether tour=pga alone fully covers co-sanctioned Euro Tour weeks.
  - Whether win/top_5/top_10/top_20/matchups odds actually post
    simultaneously or staggered (mitigated, not resolved, by checking every
    consumed market before declaring "ready" — see check_all_markets_ready).
  - softmax temperature, L1 percentile weighting, and pick-tier assignment
    thresholds are first-pass constants pending tuning against real output.

Usage:
  python weekly_model_pipeline.py            Normal unattended run.
  python weekly_model_pipeline.py --test     Detect events and log what
                                              would happen at each step,
                                              without writing files, calling
                                              the Claude API, sending email,
                                              or touching state.json.
  python weekly_model_pipeline.py --event <slug>
                                              Process only one tournament
                                              slug this firing (debugging a
                                              single event without waiting
                                              for the full detection pass).
"""

import argparse
import bisect
import csv
import json
import logging
import math
import os
import re
import smtplib
import statistics
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import matplotlib
matplotlib.use("Agg")  # non-interactive backend — required for an unattended/headless script
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────
# Paths & constants
# ─────────────────────────────────────────────────────────────────────────
REPO_ROOT = Path(__file__).resolve().parent
LOG_FILE = REPO_ROOT / "weekly_model_pipeline.log"
LOCK_FILE = REPO_ROOT / ".weekly_model_pipeline.lock"

# Firings are 3 hours apart. A run that pulls the full season of round-level
# data plus builds a 7-sheet workbook should finish in well under this, but
# give real margin before treating a lock as orphaned/stale.
LOCK_STALE_SECONDS = 30 * 60

DATAGOLF_BASE = "https://feeds.datagolf.com"
# Data Golf's documented limit is 45 requests/minute. A fixed minimum
# interval between calls is simpler and safer than a rolling counter for a
# single-threaded sequential script — 1.4s/call caps us at ~42/min.
DATAGOLF_MIN_INTERVAL = 1.4
_last_dg_call_monotonic = 0.0

TOUR_MAIN = "pga"
TOUR_OPPOSITE = "opp"

# Team/exhibition events and other tours that occasionally show up in a
# tour=pga schedule feed but aren't individual stroke-play events this
# model applies to. Duplicated from weekly_course_update.py's SKIP_KEYWORDS
# rather than imported — that script scrapes a different source
# (pgatour.com) and the two pipelines are deliberately not coupled.
SKIP_KEYWORDS = (
    "presidents cup", "ryder cup", "korn ferry", "champions tour",
    "lpga", "qbe shootout", "liv golf",
)

# Standing default per the SG Methodology doc: 60% DG skill-ratings baseline,
# 40% computed L30. Read from env vars (not just hardcoded) so a specific
# week's run can override the ratio — e.g. a FedEx Cup Playoffs week's
# smaller, hotter field weighting recent form more heavily — without editing
# this file's checked-in default. Must be set (if overriding) BEFORE this
# module is imported: FACTOR_CATALOG's labels below are built once at import
# time from these values, so a post-import monkeypatch would leave the
# labels stale even though the actual blended() math picked up the change.
SG_BLEND_SKILL_WEIGHT = float(os.environ.get("SG_BLEND_SKILL_WEIGHT_OVERRIDE", "0.60"))
SG_BLEND_L30_WEIGHT = float(os.environ.get("SG_BLEND_L30_WEIGHT_OVERRIDE", "0.40"))


def _blend_label(stat_name: str) -> str:
    skill_pct = round(SG_BLEND_SKILL_WEIGHT * 100)
    l30_pct = round(SG_BLEND_L30_WEIGHT * 100)
    return f"{stat_name} (Blended, {skill_pct}% DG skill-ratings baseline / {l30_pct}% computed DG L30)"


# The fixed whitelist of factors this pipeline can actually compute, handed
# to Claude as the menu it must choose from and weight — it may not invent
# a key outside this list. See propose_weights().
FACTOR_CATALOG = {
    "sg_app_blend": _blend_label("SG: Approach"),
    "sg_putt_blend": _blend_label("SG: Putting"),
    "sg_arg_blend": _blend_label("SG: Around the Green"),
    "sg_ott_blend": _blend_label("SG: Off the Tee"),
    "cf_approach_comp": "Course-Fit Approach Component (Data Golf player-decompositions)",
    "cf_short_comp": "Course-Fit Short-Game Component (Data Golf player-decompositions)",
    "driving_accuracy_adjustment": "Driving Accuracy Fit (Data Golf player-decompositions)",
    "driving_distance_adjustment": "Driving Distance Fit (Data Golf player-decompositions)",
    "course_history_adjustment": "Course History (Data Golf player-decompositions)",
    "course_experience_adjustment": "Course Experience (Data Golf player-decompositions)",
    "major_adjustment": "Major Championship Factor (Data Golf player-decompositions — majors only)",
    "bob_pct": "Birdie or Better % (computed from historical-raw-data/rounds)",
    "dba_pct": "Double Bogey Avoidance % — share of holes NOT double-bogey-or-worse (computed from historical-raw-data/rounds)",
    "gir_pct": "Greens in Regulation % (computed from historical-raw-data/rounds)",
    "scrambling_pct": "Scrambling % — up-and-down rate from missed greens (computed from historical-raw-data/rounds)",
    "prox_100_150_fw": "Proximity 100-150yd Fairway, SG-per-shot (Data Golf approach-skill)",
    "prox_150_200_fw": "Proximity 150-200yd Fairway, SG-per-shot (Data Golf approach-skill)",
    "rough_recovery_over150": "Rough Recovery, >150yd, SG-per-shot (Data Golf approach-skill)",
}

# Par-3 Scoring Average and Par-4/5 Scoring Average are deliberately NOT in
# this catalog. Checked directly against Data Golf's API while building
# this: historical-raw-data/rounds is round-level only (birdies/bogies/
# pars/gir/scrambling as round TOTALS, no per-hole breakdown at all), and
# historical-event-data/events has only points/earnings/finish position —
# no hole-level data exists anywhere in the historical API. The only
# hole-level endpoint (preds/live-hole-stats) is live/in-tournament only,
# not a historical archive. This is a permanent structural gap, not an
# oversight — the only way to get these two stats is a PGA Tour CSV pull
# (pgatour.com does publish them), same as the historical Scottish Open
# workbook actually sourced them from.

# ─────────────────────────────────────────────────────────────────────────
# Logging (mirrors weekly_course_update.py)
# ─────────────────────────────────────────────────────────────────────────
logger = logging.getLogger("weekly_model_pipeline")
logger.setLevel(logging.INFO)
_file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
_file_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
logger.addHandler(_file_handler)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(errors="replace")
_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(logging.Formatter("%(message)s"))
logger.addHandler(_console_handler)


# ─────────────────────────────────────────────────────────────────────────
# Concurrency guard (mirrors weekly_course_update.py)
# ─────────────────────────────────────────────────────────────────────────
class AlreadyRunningError(Exception):
    pass


def acquire_lock() -> None:
    if LOCK_FILE.exists():
        age = time.time() - LOCK_FILE.stat().st_mtime
        if age < LOCK_STALE_SECONDS:
            raise AlreadyRunningError(f"Lock file is {age:.0f}s old — another instance appears to be running.")
        logger.info(f"Removing stale lock file (age {age:.0f}s, over {LOCK_STALE_SECONDS}s threshold)")
        LOCK_FILE.unlink()
    try:
        fd = os.open(LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError:
        raise AlreadyRunningError("Another instance won the lock first.")
    os.write(fd, str(os.getpid()).encode())
    os.close(fd)


def release_lock() -> None:
    LOCK_FILE.unlink(missing_ok=True)


# ─────────────────────────────────────────────────────────────────────────
# Data Golf API — HTTP helper + rate limiting
# ─────────────────────────────────────────────────────────────────────────
class DataGolfUnavailable(Exception):
    """Raised for any failure talking to a Data Golf endpoint — network
    error, non-2xx response, or a non-JSON body. Some endpoints (confirmed:
    get-schedule?tour=opp) return a plain-text error string instead of JSON
    on an invalid/unsupported query, which json.loads() cannot parse — that
    case is folded into this exception too rather than crashing the run."""


def dg_get(path: str, params: dict) -> dict:
    global _last_dg_call_monotonic
    api_key = os.environ.get("DATAGOLF_API_KEY")
    if not api_key:
        raise DataGolfUnavailable("DATAGOLF_API_KEY environment variable is not set")

    elapsed = time.monotonic() - _last_dg_call_monotonic
    if elapsed < DATAGOLF_MIN_INTERVAL:
        time.sleep(DATAGOLF_MIN_INTERVAL - elapsed)

    query = dict(params)
    query["key"] = api_key
    query.setdefault("file_format", "json")
    url = f"{DATAGOLF_BASE}/{path}?{urllib.parse.urlencode(query)}"
    req = urllib.request.Request(url, headers={"User-Agent": "StrokesEdge-WeeklyModel/1.0"})

    _last_dg_call_monotonic = time.monotonic()
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
    except urllib.error.URLError as e:
        raise DataGolfUnavailable(f"{path}: request failed ({e})")

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        # Confirmed behavior for at least one case (get-schedule?tour=opp
        # returns "opp is not a tour we have a schedule for..." as plain
        # text). Treat any non-JSON body the same way rather than crashing.
        raise DataGolfUnavailable(f"{path}: non-JSON response ({raw[:200]!r})")


def dg_get_optional(path: str, params: dict, context: str) -> dict | None:
    """Same as dg_get but failures are logged and swallowed, returning None.
    Used specifically for tour=opp calls, where failure is the EXPECTED
    outcome on a normal single-event week — not every week has a
    concurrent opposite-field event. (Open question, noted in CLAUDE.md:
    this can't currently distinguish "no event this week" from a transient
    API problem — logged here so the first several real weeks can be
    sanity-checked by hand.)"""
    try:
        return dg_get(path, params)
    except DataGolfUnavailable as e:
        logger.info(f"{context}: {e}")
        return None


# ─────────────────────────────────────────────────────────────────────────
# Data Golf API — endpoint wrappers (all verified working this session)
# ─────────────────────────────────────────────────────────────────────────
def dg_get_schedule(tour: str = TOUR_MAIN) -> dict:
    return dg_get("get-schedule", {"tour": tour})


def dg_field_updates(tour: str) -> dict | None:
    return dg_get_optional("field-updates", {"tour": tour}, f"field-updates(tour={tour})")


def dg_skill_ratings() -> dict:
    """Tour-wide snapshot, not event-specific — fetch once per firing and
    reuse across every event being processed that run."""
    return dg_get("preds/skill-ratings", {"display": "value"})


def dg_approach_skill(period: str = "l24") -> dict:
    """period: l24 (default), l12, or ytd — no native L30/L6mo window.
    Tour-wide, fetch once per firing."""
    return dg_get("preds/approach-skill", {"period": period})


def dg_player_decompositions(tour: str) -> dict | None:
    """Event-specific (pre-tournament course-fit for whatever event is
    live under this tour code) — call per event with that event's tour
    ('pga' or 'opp')."""
    return dg_get_optional("preds/player-decompositions", {"tour": tour},
                            f"player-decompositions(tour={tour})")


def dg_historical_rounds(tour: str, year: int) -> dict:
    """Full season of round-level data in one call — large response,
    cache per (tour, year) for the duration of a firing rather than
    re-fetching per player or per event."""
    return dg_get("historical-raw-data/rounds", {"tour": tour, "event_id": "all", "year": year})


def dg_outrights(tour: str, market: str) -> dict | None:
    return dg_get_optional("betting-tools/outrights",
                            {"tour": tour, "market": market, "odds_format": "american"},
                            f"outrights(tour={tour}, market={market})")


def dg_matchups(tour: str, market: str = "tournament_matchups") -> dict | None:
    return dg_get_optional("betting-tools/matchups",
                            {"tour": tour, "market": market, "odds_format": "american"},
                            f"matchups(tour={tour}, market={market})")


def dg_fantasy_projections(tour: str, site: str = "draftkings", slate: str = "main") -> dict | None:
    """Pre-tournament DFS salaries + point/ownership projections for the
    CURRENT week. This is NOT historical-dfs-data (that endpoint only
    covers completed events, keyed by finish position, and 404s for an
    event that hasn't been played yet — confirmed by testing both against
    the live 2026 3M Open before this was wired in). preds/fantasy-
    projection-defaults is a different, separate endpoint that returns
    real, current DraftKings salaries per player (dg_id-keyed) days ahead
    of the tournament, plus a 'value' field (projected points per $1,000
    salary) DK bettors use directly. Best-effort: DK sometimes finalizes
    salaries a bit later in the week than odds/matchups do, so absence
    here must never block the rest of the pipeline."""
    return dg_get_optional("preds/fantasy-projection-defaults",
                            {"tour": tour, "site": site, "slate": slate},
                            f"fantasy-projection-defaults(tour={tour}, site={site}, slate={slate})")


# ─────────────────────────────────────────────────────────────────────────
# Slugify (duplicated from weekly_course_update.py's version — same
# behavior, deliberately not imported/coupled across the two pipelines)
# ─────────────────────────────────────────────────────────────────────────
def slugify(name: str) -> str:
    slug = name.strip().lower()
    slug = re.sub(r"[^a-z0-9\s-]", "", slug)
    slug = re.sub(r"\s+", "-", slug)
    slug = re.sub(r"-{2,}", "-", slug).strip("-")
    return slug


def is_skippable(name: str) -> bool:
    n = name.lower()
    return any(kw in n for kw in SKIP_KEYWORDS)


# ─────────────────────────────────────────────────────────────────────────
# Concurrent-event detection
# ─────────────────────────────────────────────────────────────────────────
def _make_event(entry: dict, tour: str, is_main: bool, event_type: str) -> dict:
    return {
        "event_id": entry.get("event_id"),
        "event_name": entry["event_name"],
        "course_name": entry.get("course"),
        "location": entry.get("location"),
        "start_date": entry.get("start_date"),
        "latitude": entry.get("latitude"),
        "longitude": entry.get("longitude"),
        "tour": tour,
        "is_main_event": is_main,
        "event_type": event_type,
        "slug": slugify(entry["event_name"]),
    }


def detect_events_this_week() -> list:
    """Returns one dict per PGA Tour event happening this week — always at
    least the main event, plus an opposite-field event if one exists.

    Detection logic (verified this session against a real live example —
    The Open Championship + Corales Puntacana Championship, same
    start_date, both listed under tour=pga's own schedule):
      1. get-schedule(tour=pga) already lists every event for a marquee
         week together, tagged tour="pga" regardless of which one is
         "featured" vs "opposite-field" — the schedule feed alone can't
         tell them apart.
      2. field-updates(tour=pga) is authoritative for which one is the
         main/featured event: it returns that event specifically.
      3. field-updates(tour=opp), when it succeeds, returns the opposite-
         field companion event specifically. Confirmed: on a query for
         the same week, tour=pga returned The Open Championship and
         tour=opp returned Corales Puntacana Championship — correctly
         isolated, not the same event twice.
      4. get-schedule does NOT accept tour=opp (tested — returns a
         non-JSON error string). It's only field-updates/betting-tools
         where tour=opp is a valid selector.
    """
    schedule = dg_get_schedule(TOUR_MAIN).get("schedule", [])
    upcoming = [t for t in schedule
                if str(t.get("status", "")).lower() == "upcoming"
                and t.get("event_name") and not is_skippable(t["event_name"])]
    if not upcoming:
        raise RuntimeError("No upcoming, coverable PGA Tour events found in Data Golf schedule")

    upcoming.sort(key=lambda t: t["start_date"])
    this_week_date = upcoming[0]["start_date"]
    this_week = [t for t in upcoming if t["start_date"] == this_week_date]

    events = []
    main_field = dg_field_updates(TOUR_MAIN)
    main_name = main_field.get("event_name") if main_field else None
    main_entry = next((t for t in this_week if t.get("event_name") == main_name), this_week[0])
    if main_name is None:
        logger.info("field-updates(tour=pga) returned nothing usable — "
                     f"falling back to the earliest scheduled event ('{main_entry['event_name']}') as main")
    events.append(_make_event(main_entry, tour=TOUR_MAIN, is_main=True, event_type="featured"))

    if len(this_week) > 1:
        opp_field = dg_field_updates(TOUR_OPPOSITE)
        opp_name = opp_field.get("event_name") if opp_field else None
        if opp_name:
            opp_entry = next((t for t in this_week if t.get("event_name") == opp_name), None)
            if opp_entry:
                events.append(_make_event(opp_entry, tour=TOUR_OPPOSITE, is_main=False, event_type="opposite_field"))
            else:
                logger.info(f"tour=opp resolved to '{opp_name}', which isn't among this week's "
                            f"tour=pga schedule entries — skipping it, needs manual review")
        else:
            logger.info(f"Schedule shows {len(this_week)} events this week "
                        f"({', '.join(t['event_name'] for t in this_week)}) but tour=opp returned "
                        f"nothing — building only the main event")

    for e in events:
        logger.info(f"Detected event: {e['event_name']} (tour={e['tour']}, "
                    f"is_main_event={e['is_main_event']}, slug={e['slug']})")
    return events


# ─────────────────────────────────────────────────────────────────────────
# Per-tournament directories & state
# ─────────────────────────────────────────────────────────────────────────
def event_dir(slug: str) -> Path:
    d = REPO_ROOT / slug
    d.mkdir(parents=True, exist_ok=True)
    return d


def state_path(slug: str) -> Path:
    return event_dir(slug) / "state.json"


def load_state(event: dict) -> dict:
    p = state_path(event["slug"])
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return {
        "slug": event["slug"],
        "event_name": event["event_name"],
        "tour": event["tour"],
        "is_main_event": event["is_main_event"],
        "event_type": event["event_type"],
        "start_date": event["start_date"],
        "step": "new",
        "created": datetime.now().isoformat(),
        "escalation_sent": False,
    }


def save_state(event: dict, state: dict) -> None:
    state_path(event["slug"]).write_text(json.dumps(state, indent=2), encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────
# Course characteristics — Wikipedia fallback, same pattern and same
# "mark TBD, never invent" discipline as weekly_course_update.py's
# wikipedia_par_yardage(). Duplicated rather than imported (separate
# pipeline, not coupled).
# ─────────────────────────────────────────────────────────────────────────
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"


def http_get(url: str, timeout: int = 20) -> str:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8", errors="replace")


def wikipedia_resolve_title(query: str) -> str | None:
    url = f"https://en.wikipedia.org/w/api.php?action=opensearch&search={urllib.parse.quote(query)}&limit=1&format=json"
    try:
        result = json.loads(http_get(url, timeout=10))
        titles = result[1]
        return titles[0] if titles else None
    except Exception as e:
        logger.info(f"Wikipedia title resolution failed for '{query}': {e}")
        return None


def pgatour_course_par_yardage(event_name: str) -> tuple:
    """Returns (par, yardage) as (str, int) from pgatour.com's own
    course-stats page for this event, or (None, None) if unavailable.
    Tried before the Wikipedia infobox below: it's the tour's own
    current-year figure, so it reflects mid-year course changes (a
    renovation converting par-5s to par-4s, for example) that a Wikipedia
    infobox can lag behind. Same duplicated-not-imported pattern as the
    rest of this section — self-contained, doesn't touch pgatour.com's
    /schedule the way weekly_course_update.py does, does its own lookup."""
    try:
        html = http_get("https://www.pgatour.com/schedule", timeout=20)
        m = re.search(r'<script id="__NEXT_DATA__" type="application/json">(.+?)</script>', html, re.S)
        if not m:
            return None, None
        data = json.loads(m.group(1))
        queries = data["props"]["pageProps"]["dehydratedState"]["queries"]
        schedule_q = next(q for q in queries if q["queryKey"][0] == "schedule")
        tournaments = schedule_q["state"]["data"]["tournaments"]
        match = next((t for t in tournaments if t.get("name", "").lower() == event_name.lower()), None)
        if not match:
            return None, None
        tid, year, name = match.get("tournamentId"), match.get("year"), match.get("name")
        if not tid or not year or not name:
            return None, None
        slug = slugify(name)
        stats_html = http_get(f"https://www.pgatour.com/tournaments/{year}/{slug}/{tid}/course-stats", timeout=15)
        m2 = re.search(r'<script id="__NEXT_DATA__" type="application/json">(.+?)</script>', stats_html, re.S)
        if not m2:
            return None, None
        stats_data = json.loads(m2.group(1))
        # Target the "courseStats" query specifically, not just any query
        # with a truthy "courses" list — this page's __NEXT_DATA__ bundles
        # queries for every tour's tournament running the same week (e.g.
        # PGA Tour + Korn Ferry + Champions + DP World, each keyed by its
        # own tournamentId), and several of THOSE "tournament" queries also
        # carry a "courses" list, just a bare id/courseName one with no
        # par/yardage fields. A bare `next(... and data.get("courses"))`
        # matches the first of those (often even the right tournament ID,
        # just the wrong query shape) and returns par=None/yardage=None
        # before ever reaching the "courseStats" query that actually has
        # them — confirmed live 2026-08-03: the Wyndham Championship page's
        # first "tournament" query WAS R2026013 (correct event) but its
        # courses entry only had id/courseName/courseCode, silently
        # producing a (None, None) result and falling back to a stale
        # Wikipedia yardage (7127 vs the real 7131) even though the
        # correct par/yardage were sitting a few queries later.
        course_q = next((q for q in stats_data["props"]["pageProps"]["dehydratedState"]["queries"]
                         if q.get("queryKey", [None])[0] == "courseStats"
                         and isinstance(q.get("state", {}).get("data"), dict)
                         and q["state"]["data"].get("courses")), None)
        if not course_q:
            return None, None
        course = course_q["state"]["data"]["courses"][0]
        par, yardage = course.get("par"), course.get("yardage")
        if not par or not yardage:
            return None, None
        return str(par), int(yardage)
    except Exception as e:
        logger.info(f"pgatour.com course-stats fetch failed for '{event_name}': {e}")
        return None, None


def _clean_wikitext_value(raw: str) -> str:
    """Strip wikilinks/refs/line-breaks down to plain text, e.g.
    '[[Arnold Palmer]],<br>with [[Tom Lehman]]' -> 'Arnold Palmer, with Tom Lehman'
    and multi-name '<br>'-separated lists (no source comma) -> comma-joined."""
    s = re.sub(r"<ref[^>]*/?>.*?(</ref>|$)", "", raw, flags=re.I | re.S)
    s = re.sub(r"<br\s*/?>", ", ", s, flags=re.I)
    s = re.sub(r"\[\[(?:[^\|\]]+\|)?([^\]]+)\]\]", r"\1", s)  # [[Target|Display]] -> Display, [[Target]] -> Target
    s = re.sub(r"<[^>]+>", "", s)
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"(,\s*){2,}", ", ", s)  # source already-comma'd names + our own <br>-comma can double up
    s = s.strip().strip(",").strip()
    return s


def wikipedia_course_facts(course_name: str) -> dict:
    """Best-effort only. Missing fields are marked None/"TBD" rather than
    guessed — the weight-proposal prompt is told explicitly which facts
    are unavailable so Claude doesn't fabricate specifics from a gap.

    Includes designer/established-year alongside par/yardage/course_type —
    added after a real proposal (3M Open) stated a fabricated architect
    ("Rees Jones") for TPC Twin Cities (actually Arnold Palmer/Tom Lehman).
    Root cause: the weight-proposal prompt never sourced architect/course-
    history facts from anywhere, so Claude filled that gap from its own
    unverified training recall. Same infobox already parsed for par/yardage
    also carries a designer1/designer field, so this closes the gap with
    no new data source — same "known -> state it, unknown -> don't invent"
    discipline as par/yardage/course_type above."""
    facts = {"par": None, "yardage": None, "course_type": "unknown",
             "designer": None, "established_year": None}
    title = wikipedia_resolve_title(course_name)
    if not title:
        return facts
    try:
        url = (f"https://en.wikipedia.org/w/api.php?action=parse&page="
               f"{urllib.parse.quote(title)}&prop=wikitext&format=json")
        data = json.loads(http_get(url, timeout=10))
        wikitext = data["parse"]["wikitext"]["*"]
    except Exception as e:
        logger.info(f"Wikipedia wikitext fetch failed for '{title}': {e}")
        return facts

    infobox = re.search(r"\{\{Infobox golf[^\n]*\n(.*?)\n\}\}", wikitext, re.S | re.I)
    if infobox:
        body = infobox.group(1)
        par_m = re.search(r"\bpar1?\s*=\s*([0-9/]+)", body, re.I)
        len_m = re.search(r"\blength1?\s*=\s*\{\{convert\|([0-9,]+)\|yd", body, re.I)
        facts["par"] = par_m.group(1).split("/")[0] if par_m else None
        yards = len_m.group(1).replace(",", "") if len_m else None
        facts["yardage"] = int(yards) if yards else None

        designer_m = re.search(r"\bdesigner1?\s*=\s*(.+)", body, re.I)
        if designer_m:
            cleaned = _clean_wikitext_value(designer_m.group(1))
            facts["designer"] = cleaned or None

        established_m = re.search(r"\bestablished\s*=\s*(.+)", body, re.I)
        if established_m:
            year_m = re.search(r"(\d{4})", established_m.group(1))
            facts["established_year"] = int(year_m.group(1)) if year_m else None

    # Strip the boilerplate tail before keyword-checking course_type.
    # "External links" is a standard section heading on nearly every
    # Wikipedia article, and the naive "links" substring check below was
    # matching on that heading alone, not real prose — confirmed live
    # 2026-08-03: Wyndham Championship, Sedgefield Country Club, AND
    # Detroit Golf Club's articles all contain exactly one "links" match,
    # every time from "==External links==", regardless of actual course
    # type. This silently mislabeled Detroit Golf Club (a parkland
    # course) as "links" last week too — state.json shows "parkland" now
    # only because it was hand-corrected after the fact, not because this
    # function got it right. Only the article's real body prose (and
    # infobox, already parsed above) should ever feed this heuristic.
    body_text = re.split(r"==\s*(?:External links|References|See also)\s*==", wikitext, flags=re.I)[0]
    text_lower = body_text.lower()
    # Broad "links" check first (a Royal Birkdale-style article often just
    # says "links" without "links course" as an exact phrase — confirmed
    # by running this against the real article while testing), narrower
    # phrases as a fallback for articles that phrase it differently.
    if "links" in text_lower or any(k in text_lower for k in ("seaside", "dunes")):
        facts["course_type"] = "links"
    elif any(k in text_lower for k in ("desert course", "scottsdale", "palm springs")):
        facts["course_type"] = "desert"
    else:
        facts["course_type"] = "parkland"
    return facts


# ─────────────────────────────────────────────────────────────────────────
# Weight proposal — Claude API call (unattended, same env-var pattern as
# weekly_course_update.py's generate_analysis()). Deliberately claude-
# sonnet-5, not haiku: a mispriced L2 gate is a more consequential mistake
# than imprecise course-profile prose, so this call gets the stronger
# model even though it costs more per run.
# ─────────────────────────────────────────────────────────────────────────
CLAUDE_MODEL_WEIGHTS = "claude-sonnet-5"
CLAUDE_MAX_TOKENS_WEIGHTS = 2500

WEIGHT_PROPOSAL_PROMPT = """You are the model architect for StrokesEdge, a quantitative PGA Tour golf betting brand. Propose Layer-1 course-fit weights and Layer-2 hard gates for this week's model.

TOURNAMENT: {event_name}
COURSE: {course_name}
LOCATION: {location}
PAR: {par}
YARDAGE: {yardage}
COURSE TYPE (best-effort inference, may be wrong): {course_type}
ARCHITECT / DESIGNER (verified from Wikipedia infobox): {designer}
YEAR OPENED (verified from Wikipedia infobox): {established_year}
IS THIS A MAJOR CHAMPIONSHIP: {is_major}

Any field above marked "unknown" or "None" was genuinely unavailable from the data source used to build this prompt — do not invent a specific number or name for it. It is fine to reason qualitatively about a course you don't have exact par/yardage for.

ARCHITECT/HISTORY DISCIPLINE: only state the architect/designer if it is given above verbatim (never guess or recall one from memory, even if you believe you know it — course architect attributions are exactly the kind of specific fact that's easy to misremember and this has produced a real factual error before). Same for year opened. Any other course-history claim (notable past champions, prior tournament names, renovations, etc.) must stay qualitative/generic unless it's a widely-established fact you're highly confident of — when in doubt, omit it rather than state a specific name, date, or number you're not certain of.

AVAILABLE FACTORS — you may ONLY use these exact keys, you may not invent new ones:
{factor_catalog}

IMPORTANT SCALE NOTE: every factor above is sourced from Data Golf, not PGA Tour's own stats. Data Golf's "True SG" numbers run on a different scale than PGA Tour's official SG stat (verified: some players differ by nearly 2x between the two sources). Calibrate your L2 gate thresholds to what's plausible for DATA GOLF's scale specifically — do not reuse PGA-Tour-calibrated thresholds you might have seen elsewhere (e.g. a Data-Golf-scale approach gate is more likely to look like ">= -0.10" than ">= +0.20").

COURSE CHANGE CHECK: material physical changes to a course (a restoration, a redesign, holes rebuilt or converted between par values, greens rebuilt) make Course History and Course Experience weight less than usual, since most of the field's past results/reps happened on a course that no longer exists in that form. Only state yes/uncertain if this is something you're genuinely confident about for THIS specific course, not a guess extended from "old courses sometimes get renovated" — when in doubt, say no, same discipline as the architect/history rule above.

TASK — respond with ONLY the following, no other prose, no markdown headers beyond what's shown, filled in exactly:

COURSE_CHANGE_FLAG: <yes / no / uncertain> — <if yes or uncertain: one sentence on what changed and that Course History/Course Experience weights below should get manual review before approval; if no: "No indication of a material recent course change.">

SUMMARY: <2-4 sentences: what this course demands and why the weights below reflect that>

WEIGHTS:
- <factor_key> | <weight_pct, a plain number, no % sign> | <one-sentence rationale citing something about THIS course>
(repeat one line per factor you're using — weights must sum to exactly 100)

GATES:
- <factor_key> | <operator, one of these four: gte, lte, gt, lt> | <threshold value, a plain number> | <one-sentence rationale>
(2 to 4 gates — the hard thresholds every plausible winner at this course should clear)
"""


def propose_weights(event: dict, course_facts: dict, is_major: bool) -> str:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY environment variable is not set")

    catalog_text = "\n".join(f"  {k} — {v}" for k, v in FACTOR_CATALOG.items())
    prompt = WEIGHT_PROPOSAL_PROMPT.format(
        event_name=event["event_name"],
        course_name=event.get("course_name") or "unknown",
        location=event.get("location") or "unknown",
        par=course_facts.get("par") or "unknown",
        yardage=course_facts.get("yardage") or "unknown",
        course_type=course_facts.get("course_type") or "unknown",
        designer=course_facts.get("designer") or "unknown",
        established_year=course_facts.get("established_year") or "unknown",
        is_major="yes" if is_major else "no",
        factor_catalog=catalog_text,
    )
    body = json.dumps({
        "model": CLAUDE_MODEL_WEIGHTS,
        "max_tokens": CLAUDE_MAX_TOKENS_WEIGHTS,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=body,
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=90) as resp:
        result = json.loads(resp.read().decode("utf-8"))
    return "".join(block.get("text", "") for block in result.get("content", []))


# ─────────────────────────────────────────────────────────────────────────
# weights_proposal.md — parse Claude's raw text into structured
# weights/gates, render the human-editable file, and re-parse it later to
# pick up any edits Brian made before approving.
# ─────────────────────────────────────────────────────────────────────────
OPERATOR_MAP = {"gte": ">=", "lte": "<=", "gt": ">", "lt": "<"}
OPERATOR_FUNCS = {
    ">=": lambda a, b: a >= b, "<=": lambda a, b: a <= b,
    ">": lambda a, b: a > b, "<": lambda a, b: a < b,
}


def parse_claude_weight_response(raw_text: str) -> dict:
    """Parses the SUMMARY/WEIGHTS/GATES block format from the prompt above.
    Raises ValueError with a clear message if Claude didn't follow the
    format — caller should treat that as a failed proposal, not silently
    proceed with partial data."""
    flag_m = re.search(r"COURSE_CHANGE_FLAG:\s*(yes|no|uncertain)\s*—\s*(.+?)(?=\nSUMMARY:)", raw_text, re.S | re.I)
    if flag_m:
        course_change_flag = {"status": flag_m.group(1).strip().lower(), "note": flag_m.group(2).strip()}
    else:
        # Model didn't return this field in the expected shape — treat as
        # "uncertain" rather than silently assuming "no", so a parsing miss
        # still surfaces for a human to check rather than disappearing.
        course_change_flag = {"status": "uncertain", "note": "Model did not return a parseable COURSE_CHANGE_FLAG line — check manually."}

    summary_m = re.search(r"SUMMARY:\s*(.+?)(?=\nWEIGHTS:)", raw_text, re.S)
    summary = summary_m.group(1).strip() if summary_m else "(no summary returned)"

    weights = []
    weights_block_m = re.search(r"WEIGHTS:\s*(.+?)(?=\nGATES:)", raw_text, re.S)
    if weights_block_m:
        for line in weights_block_m.group(1).splitlines():
            m = re.match(r"\s*-\s*([a-z0-9_]+)\s*\|\s*([\d.]+)\s*\|\s*(.+)", line)
            if m:
                key, pct, rationale = m.groups()
                if key not in FACTOR_CATALOG:
                    logger.info(f"Claude proposed unknown factor key '{key}' — dropping it")
                    continue
                weights.append({"key": key, "weight_pct": float(pct), "rationale": rationale.strip()})
    if not weights:
        raise ValueError("Could not parse any WEIGHTS lines from Claude's response")

    total = sum(w["weight_pct"] for w in weights)
    if abs(total - 100.0) > 0.5:
        logger.info(f"Proposed weights sum to {total}, not 100 — rescaling proportionally")
        for w in weights:
            w["weight_pct"] = round(w["weight_pct"] * 100.0 / total, 2)

    gates = []
    gates_block_m = re.search(r"GATES:\s*(.+)", raw_text, re.S)
    if gates_block_m:
        for line in gates_block_m.group(1).splitlines():
            m = re.match(r"\s*-\s*([a-z0-9_]+)\s*\|\s*(gte|lte|gt|lt)\s*\|\s*(-?[\d.]+)\s*\|\s*(.+)", line)
            if m:
                key, op_word, val, rationale = m.groups()
                if key not in FACTOR_CATALOG:
                    logger.info(f"Claude proposed a gate on unknown factor key '{key}' — dropping it")
                    continue
                gates.append({"key": key, "operator": OPERATOR_MAP[op_word], "value": float(val),
                              "rationale": rationale.strip()})
    if not gates:
        raise ValueError("Could not parse any GATES lines from Claude's response")

    return {"summary": summary, "weights": weights, "gates": gates, "course_change_flag": course_change_flag}


def weights_proposal_path(slug: str) -> Path:
    return event_dir(slug) / "weights_proposal.md"


# The 4 SG categories present in EVERY historical workbook reviewed while
# building this feature, without exception, regardless of course type —
# the closest thing to a "universal" factor this model has. Backtesting
# against 4 real historical workbooks found the fresh proposal dropping
# one of these to zero in 3 of 4 cases (most notably SG:OTT at 18% in a
# human-built ISCO Championship model vs. 0% here) — a real, repeat
# disagreement worth flagging, not the long-tail contextual factors
# (proximity bands, course history, etc.) which are legitimately optional
# per course and zero out routinely for good reason.
CORE_SG_FACTORS = ("sg_app_blend", "sg_putt_blend", "sg_arg_blend", "sg_ott_blend")


def detect_zeroed_core_factors(proposal: dict) -> list:
    """Returns a list of human-readable warnings, one per core SG category
    absent from the proposed weights (Claude's response format only lists
    factors it chose to weight — an absent key IS a zero, just an implicit
    one, easy to miss on a skim). Advisory only — does not block approval,
    per the ask: make it impossible to miss, not impossible to accept."""
    used_keys = {w["key"] for w in proposal["weights"]}
    warnings = []
    for key in CORE_SG_FACTORS:
        if key not in used_keys:
            label = FACTOR_CATALOG.get(key, key)
            warnings.append(f"{key} ({label}) — ZEROED, not included in this week's weights at all")
    return warnings


def render_proposal_md(event: dict, proposal: dict, status: str = "PENDING REVIEW") -> str:
    zero_warnings = detect_zeroed_core_factors(proposal)
    lines = [
        f"STATUS: {status}",
        "",
        f"# StrokesEdge Weekly Model — Weight Proposal",
        f"Tournament: {event['event_name']}",
        f"Course: {event.get('course_name') or 'unknown'}",
        f"Event type: {event['event_type']} ({'main/featured event' if event['is_main_event'] else 'opposite-field event'})",
        f"Generated: {datetime.now().isoformat(timespec='minutes')}",
    ]
    if zero_warnings:
        lines += ["", "!! ZEROED CORE FACTORS — CONFIRM THIS IS INTENTIONAL BEFORE APPROVING !!"]
        for w in zero_warnings:
            lines.append(f"  - {w}")
    change_flag = proposal.get("course_change_flag", {})
    if change_flag.get("status") in ("yes", "uncertain"):
        lines += ["", "!! COURSE CHANGE FLAG — REVIEW COURSE HISTORY / COURSE EXPERIENCE WEIGHTS BEFORE APPROVING !!",
                  f"  - {change_flag['note']}"]
    lines += [
        "",
        "## Course Summary",
        proposal["summary"],
        "",
        "## L1 Weights (must sum to 100)",
        "# format: - key | weight_pct | rationale",
    ]
    for w in proposal["weights"]:
        label = FACTOR_CATALOG.get(w["key"], w["key"])
        lines.append(f"- {w['key']} | {w['weight_pct']:g} | {w['rationale']}   [{label}]")
    lines += ["", "## L2 Gates", "# format: - key | operator | value | rationale  (operator: >= <= > <)"]
    for g in proposal["gates"]:
        label = FACTOR_CATALOG.get(g["key"], g["key"])
        lines.append(f"- {g['key']} | {g['operator']} | {g['value']:g} | {g['rationale']}   [{label}]")
    lines += [
        "",
        "## How to approve or edit",
        "Leaving everything as-is and changing STATUS above to APPROVED accepts this proposal unchanged.",
        "To change a number: edit the weight_pct or value directly on its line (keep the '- key | ... |' shape",
        "intact so it can still be read back). Weights don't need to be hand-rebalanced to sum to 100 —",
        "the pipeline rescales proportionally if they don't. Then change STATUS to APPROVED.",
        "The next scheduled firing (within a few hours) picks up whatever is in this file at that moment.",
    ]
    return "\n".join(lines) + "\n"


def write_proposal(event: dict, proposal: dict) -> Path:
    path = weights_proposal_path(event["slug"])
    path.write_text(render_proposal_md(event, proposal), encoding="utf-8")
    return path


def parse_proposal_file(path: Path) -> dict:
    """Re-parses a (possibly hand-edited) weights_proposal.md. Same line
    shape as render_proposal_md's output, tolerant of the trailing
    '[label]' hint and of the STATUS/section furniture around it."""
    text = path.read_text(encoding="utf-8")
    status_m = re.search(r"^STATUS:\s*(.+)$", text, re.M)
    status = status_m.group(1).strip().upper() if status_m else "PENDING REVIEW"

    weights = []
    for m in re.finditer(r"^-\s*([a-z0-9_]+)\s*\|\s*([\d.]+)\s*\|\s*([^\[\n]+)", text, re.M):
        key, pct, rationale = m.groups()
        weights.append({"key": key, "weight_pct": float(pct), "rationale": rationale.strip()})

    gates = []
    for m in re.finditer(r"^-\s*([a-z0-9_]+)\s*\|\s*(>=|<=|>|<)\s*\|\s*(-?[\d.]+)\s*\|\s*([^\[\n]+)", text, re.M):
        key, op, val, rationale = m.groups()
        gates.append({"key": key, "operator": op, "value": float(val), "rationale": rationale.strip()})

    return {"status": status, "weights": weights, "gates": gates}


def check_approval(event: dict) -> dict | None:
    """Returns the (possibly Brian-edited) parsed proposal if the file's
    STATUS line reads APPROVED, else None. A missing file is treated as
    not-yet-approved rather than an error — nothing to do but wait."""
    path = weights_proposal_path(event["slug"])
    if not path.exists():
        return None
    parsed = parse_proposal_file(path)
    if parsed["status"] != "APPROVED":
        return None
    if not parsed["weights"] or not parsed["gates"]:
        logger.error(f"[{event['slug']}] weights_proposal.md is marked APPROVED but failed to parse "
                     f"weights/gates from it — treating as not-yet-approved rather than proceeding with nothing")
        return None
    total = sum(w["weight_pct"] for w in parsed["weights"])
    if total > 0 and abs(total - 100.0) > 0.5:
        for w in parsed["weights"]:
            w["weight_pct"] = w["weight_pct"] * 100.0 / total
    return parsed


# ─────────────────────────────────────────────────────────────────────────
# L30 recent-form blend — computed from historical-raw-data/rounds, since
# Data Golf has no native L30 (or L6mo) query (verified this session:
# preds/approach-skill only supports period=l24|l12|ytd; skill-ratings has
# no period parameter at all). Season-of-rounds response is large — cache
# per (tour, year) for the life of one firing.
# ─────────────────────────────────────────────────────────────────────────
L30_WINDOW_DAYS = 30
L30_MIN_ROUNDS = 3  # below this, treat the player's L30 as unavailable, not a bad signal

_rounds_cache: dict = {}


def get_year_rounds_cached(tour: str, year: int) -> dict:
    key = (tour, year)
    if key not in _rounds_cache:
        logger.info(f"Fetching historical-raw-data/rounds (tour={tour}, year={year}) — large response, caching")
        _rounds_cache[key] = dg_historical_rounds(tour, year)
    return _rounds_cache[key]


def compute_l30_window_stats(tour: str, as_of: date) -> tuple:
    """Returns (l30_by_dg_id, rate_by_dg_id). l30_by_dg_id maps dg_id ->
    {'sg_app','sg_arg','sg_ott','sg_putt','rounds'}; entries with fewer
    than L30_MIN_ROUNDS are omitted entirely (caller must treat a missing
    dg_id as 'unavailable', never assume 0). rate_by_dg_id maps dg_id ->
    {'bob_pct','dba_pct','gir_pct','scrambling_pct','rounds'} over the same
    window and same minimum-sample rule — birdie-or-better, double-bogey-
    or-worse, greens in regulation, and scrambling (up-and-down from missed
    greens) are all direct per-round fields on historical-raw-data/rounds,
    aggregated the same way as the SG stats. dba_pct is the mirror of
    bob_pct: share of holes that were NOT a double bogey or worse, same
    holes-played denominator, so a higher number is always better for both.
    """
    cutoff = as_of - timedelta(days=L30_WINDOW_DAYS)
    rounds_data = get_year_rounds_cached(tour, as_of.year)

    accum: dict = {}
    for event in rounds_data.values():
        if not isinstance(event, dict):
            continue
        completed_raw = event.get("event_completed")
        if not completed_raw:
            continue
        try:
            completed = date.fromisoformat(completed_raw[:10])
        except ValueError:
            continue
        # event_completed is the only date field this endpoint returns —
        # it's event-level, not per-round, so the L30 window is
        # necessarily coarser than a true per-round 30-day filter. Noted
        # as a known granularity limitation, not a bug.
        if not (cutoff <= completed <= as_of):
            continue
        for score in event.get("scores", []):
            dg_id = score.get("dg_id")
            if dg_id is None:
                continue
            bucket = accum.setdefault(dg_id, {"sg_app": [], "sg_arg": [], "sg_ott": [], "sg_putt": [],
                                                "birdies_plus": 0, "doubles_plus": 0, "holes": 0, "rounds": 0,
                                                "gir_sum": 0.0, "scrambling_sum": 0.0, "rate_rounds": 0})
            for rk, rv in score.items():
                if not rk.startswith("round_") or not isinstance(rv, dict):
                    continue
                bucket["rounds"] += 1
                for stat in ("sg_app", "sg_arg", "sg_ott", "sg_putt"):
                    if isinstance(rv.get(stat), (int, float)):
                        bucket[stat].append(rv[stat])
                birdies = rv.get("birdies") or 0
                eagles_plus = rv.get("eagles_or_better") or 0
                bucket["birdies_plus"] += birdies + eagles_plus
                bucket["doubles_plus"] += rv.get("doubles_or_worse") or 0
                bucket["holes"] += 18  # Data Golf round records don't expose a holes-played count directly
                # gir/scrambling are per-round FRACTIONS (0-1), not counts — average them
                # across rounds rather than summing, so a round with a missing value doesn't
                # need a holes-played denominator the way birdie counting does.
                if isinstance(rv.get("gir"), (int, float)) and isinstance(rv.get("scrambling"), (int, float)):
                    bucket["gir_sum"] += rv["gir"]
                    bucket["scrambling_sum"] += rv["scrambling"]
                    bucket["rate_rounds"] += 1

    l30, rates = {}, {}
    for dg_id, b in accum.items():
        if b["rounds"] < L30_MIN_ROUNDS:
            continue
        l30[dg_id] = {
            "sg_app": sum(b["sg_app"]) / len(b["sg_app"]) if b["sg_app"] else None,
            "sg_arg": sum(b["sg_arg"]) / len(b["sg_arg"]) if b["sg_arg"] else None,
            "sg_ott": sum(b["sg_ott"]) / len(b["sg_ott"]) if b["sg_ott"] else None,
            "sg_putt": sum(b["sg_putt"]) / len(b["sg_putt"]) if b["sg_putt"] else None,
            "rounds": b["rounds"],
        }
        entry = {"rounds": b["rounds"]}
        if b["holes"] > 0:
            entry["bob_pct"] = b["birdies_plus"] / b["holes"] * 100.0
            entry["dba_pct"] = 100.0 - (b["doubles_plus"] / b["holes"] * 100.0)
        if b["rate_rounds"] > 0:
            entry["gir_pct"] = b["gir_sum"] / b["rate_rounds"] * 100.0
            entry["scrambling_pct"] = b["scrambling_sum"] / b["rate_rounds"] * 100.0
        if len(entry) > 1:  # more than just 'rounds'
            rates[dg_id] = entry
    return l30, rates


# ─────────────────────────────────────────────────────────────────────────
# PGA Tour CSV supplement — optional, checked at firing time, never
# required. Never blended into the same column as a Data-Golf-sourced
# stat (see SG Methodology in CLAUDE.md) — feeds only the categories Data
# Golf can't match at the same granularity (proximity bands, etc).
# ─────────────────────────────────────────────────────────────────────────
PGA_SUPPLEMENT_FILENAME = "pga_tour_supplement.csv"


def load_pga_supplement(event: dict) -> dict | None:
    path = event_dir(event["slug"]) / PGA_SUPPLEMENT_FILENAME
    if not path.exists():
        return None
    try:
        with path.open(encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f))
    except Exception as e:
        logger.error(f"[{event['slug']}] found {PGA_SUPPLEMENT_FILENAME} but failed to read it: {e}")
        return None
    if not rows or "Player" not in rows[0]:
        logger.error(f"[{event['slug']}] {PGA_SUPPLEMENT_FILENAME} exists but has no 'Player' column — ignoring it")
        return None
    logger.info(f"[{event['slug']}] using PGA Tour CSV supplement ({len(rows)} players)")
    return {row["Player"]: row for row in rows}


# ─────────────────────────────────────────────────────────────────────────
# Player metrics — assembles every FACTOR_CATALOG value per player from
# the various Data Golf endpoints into one flat lookup structure.
# ─────────────────────────────────────────────────────────────────────────
MAJOR_KEYWORDS = ("masters", "u.s. open", "us open", "the open championship", "pga championship")


def is_major_event(event_name: str) -> bool:
    n = event_name.lower()
    return any(k in n for k in MAJOR_KEYWORDS)


def _index_by_dg_id(items: list) -> dict:
    return {item["dg_id"]: item for item in items if "dg_id" in item}


def build_player_metrics(event: dict, field: dict, skill_ratings: dict, approach_skill: dict,
                          decompositions: dict | None, l30_by_id: dict, rate_by_id: dict) -> dict:
    """Returns dg_id -> {'player_name', 'sample_ok', 'l30_available', 'factors': {key: value|None}}.

    sample_ok is the Watch List carve-out: True only if Data Golf's own
    skill-ratings has real SG data for this player (the Data-Golf-only
    equivalent of "enough PGA Tour ShotLink rounds" from the pre-automation
    workbooks — LIV/DP World players or anyone Data Golf can't compute a
    reliable SG split for will be missing here and go to the Watch List
    sheet instead of L1/L2 regression)."""
    skill_idx = _index_by_dg_id(skill_ratings.get("players", []))
    approach_idx = _index_by_dg_id(approach_skill.get("data", []))
    decomp_idx = _index_by_dg_id(decompositions.get("players", [])) if decompositions else {}

    metrics = {}
    for entry in field.get("field", []):
        dg_id = entry.get("dg_id")
        if dg_id is None:
            continue
        skill = skill_idx.get(dg_id)
        sample_ok = bool(skill and skill.get("sg_total") is not None)

        l30 = l30_by_id.get(dg_id)
        l30_available = l30 is not None

        def blended(stat_key: str) -> float | None:
            base = skill.get(stat_key) if skill else None
            if base is None:
                return None
            if l30 and l30.get(stat_key) is not None:
                return round(SG_BLEND_SKILL_WEIGHT * base + SG_BLEND_L30_WEIGHT * l30[stat_key], 4)
            return base  # L30 unavailable — use baseline alone, flagged via l30_available

        appr = approach_idx.get(dg_id, {})
        decomp = decomp_idx.get(dg_id, {})
        rates = rate_by_id.get(dg_id)

        factors = {
            "sg_app_blend": blended("sg_app"),
            "sg_putt_blend": blended("sg_putt"),
            "sg_arg_blend": blended("sg_arg"),
            "sg_ott_blend": blended("sg_ott"),
            "cf_approach_comp": decomp.get("cf_approach_comp"),
            "cf_short_comp": decomp.get("cf_short_comp"),
            "driving_accuracy_adjustment": decomp.get("driving_accuracy_adjustment"),
            "driving_distance_adjustment": decomp.get("driving_distance_adjustment"),
            "course_history_adjustment": decomp.get("course_history_adjustment"),
            "course_experience_adjustment": decomp.get("course_experience_adjustment"),
            "major_adjustment": decomp.get("major_adjustment"),
            "bob_pct": rates.get("bob_pct") if rates else None,
            "dba_pct": rates.get("dba_pct") if rates else None,
            "gir_pct": rates.get("gir_pct") if rates else None,
            "scrambling_pct": rates.get("scrambling_pct") if rates else None,
            "prox_100_150_fw": appr.get("100_150_fw_sg_per_shot"),
            "prox_150_200_fw": appr.get("150_200_fw_sg_per_shot"),
            "rough_recovery_over150": appr.get("over_150_rgh_sg_per_shot"),
        }
        metrics[dg_id] = {
            "player_name": entry.get("player_name", f"dg_id {dg_id}"),
            "sample_ok": sample_ok,
            "l30_available": l30_available,
            "l30_rounds": l30["rounds"] if l30 else 0,
            "owgr_rank": entry.get("owgr_rank"),
            "dg_rank": entry.get("dg_rank"),
            "factors": factors,
        }
    return metrics


# ─────────────────────────────────────────────────────────────────────────
# L1 — percentile-weighted composite score
# ─────────────────────────────────────────────────────────────────────────
def _percentiles_for_factor(metrics: dict, key: str) -> dict:
    values = {dg_id: m["factors"].get(key) for dg_id, m in metrics.items()
              if m["sample_ok"] and m["factors"].get(key) is not None}
    if not values:
        return {}
    ordered = sorted(values.values())
    n = len(ordered)
    return {dg_id: bisect.bisect_right(ordered, v) / n * 100.0 for dg_id, v in values.items()}


def run_l1(metrics: dict, weights: list) -> dict:
    """Percentile-rank each player against the field for each weighted
    factor (0-100), weighted-average those percentiles into a single L1
    score. A factor missing for a given player is excluded from THAT
    player's average and the remaining weights are implicitly renormalized
    (divide by weight actually used, not by 100) — never assumed as a
    league-average value. weight_coverage_pct on the result shows how much
    of the intended weight was actually usable for that player, so a
    player scored off e.g. 60% weight coverage is visibly different from
    one scored off 100%."""
    percentile_tables = {w["key"]: _percentiles_for_factor(metrics, w["key"]) for w in weights}

    results = {}
    for dg_id, m in metrics.items():
        if not m["sample_ok"]:
            continue
        components, weighted_sum, weight_used = {}, 0.0, 0.0
        for w in weights:
            pct = percentile_tables[w["key"]].get(dg_id)
            if pct is None:
                continue
            components[w["key"]] = {"value": m["factors"][w["key"]], "percentile": round(pct, 1),
                                     "weight_pct": w["weight_pct"]}
            weighted_sum += pct * w["weight_pct"]
            weight_used += w["weight_pct"]
        l1_score = round(weighted_sum / weight_used, 1) if weight_used > 0 else None
        results[dg_id] = {"l1_score": l1_score, "components": components,
                           "weight_coverage_pct": round(weight_used, 1)}
    return results


# ─────────────────────────────────────────────────────────────────────────
# L2 — hard gates
# ─────────────────────────────────────────────────────────────────────────
def apply_l2(metrics: dict, gates: list) -> dict:
    results = {}
    for dg_id, m in metrics.items():
        if not m["sample_ok"]:
            continue
        detail, passed = [], True
        for g in gates:
            val = m["factors"].get(g["key"])
            if val is None:
                passed = False
                detail.append(f"{g['key']}: unavailable (treated as FAIL, not assumed)")
                continue
            ok = OPERATOR_FUNCS[g["operator"]](val, g["value"])
            detail.append(f"{g['key']} {g['operator']} {g['value']:g} -> "
                          f"{'PASS' if ok else 'FAIL'} (actual {_fmt_factor_value(val)})")
            passed = passed and ok
        results[dg_id] = {"pass": passed, "detail": detail}
    return results


# ─────────────────────────────────────────────────────────────────────────
# Odds / market readiness
# ─────────────────────────────────────────────────────────────────────────
def _has_real_odds(resp: dict | None) -> bool:
    return bool(resp) and isinstance(resp.get("odds"), list) and len(resp["odds"]) > 0


def _has_real_matchups(resp: dict | None) -> bool:
    return bool(resp) and isinstance(resp.get("match_list"), list) and len(resp["match_list"]) > 0


def check_markets(tour: str) -> dict:
    """Outrights (win/top_5/top_10/top_20) are the hard gate for 'ready to
    build the workbook'. Matchups are checked and used if available, but
    deliberately NOT allowed to block the whole week indefinitely — a
    smaller or alternate-field event may simply never get a meaningful
    matchup market from any book. This is a considered deviation from
    "wait for every market," not an oversight: if matchups never show up,
    the workbook still ships without a Matchup tier rather than never
    shipping at all."""
    markets = {}
    outrights_ready = True
    for market in ("win", "top_5", "top_10", "top_20"):
        resp = dg_outrights(tour, market)
        ok = _has_real_odds(resp)
        markets[f"outrights_{market}"] = {"ready": ok, "response": resp}
        outrights_ready = outrights_ready and ok
    matchup_resp = dg_matchups(tour, "tournament_matchups")
    matchups_ready = _has_real_matchups(matchup_resp)
    markets["matchups"] = {"ready": matchups_ready, "response": matchup_resp}
    return {"outrights_ready": outrights_ready, "matchups_ready": matchups_ready, "markets": markets}


# ─────────────────────────────────────────────────────────────────────────
# L3 — value screen
# ─────────────────────────────────────────────────────────────────────────
PREFERRED_BOOKS = ("fanduel", "draftkings", "betmgm", "caesars", "bet365")


def pick_book_odds(odds_row: dict) -> tuple:
    """Historical workbooks labeled this column 'ODDS (BR)' for
    BetRivers — Data Golf's tracked-books list does not include
    BetRivers at all (confirmed: bet365, betcris, betmgm, betonline,
    betway, bovada, caesars, draftkings, fanduel, pinnacle, pointsbet,
    skybet, williamhill, unibet). Brian's actual primary books are
    FanDuel and BetRivers — FanDuel is tracked (confirmed present in both
    books_offering and real odds rows), so it's the preferred reference
    book instead of BetRivers. The workbook labels this column
    'ODDS (FD)', not '(BR)', to be accurate about the actual source."""
    for book in PREFERRED_BOOKS:
        if odds_row.get(book) not in (None, "", "-"):
            return book, odds_row[book]
    for k, v in odds_row.items():
        if k not in ("dg_id", "player_name", "datagolf") and v not in (None, "", "-"):
            return k, v
    return None, None


def implied_prob_pct(odds_val) -> float | None:
    try:
        o = float(odds_val)
    except (TypeError, ValueError):
        return None
    if o >= 0:
        return 100.0 / (o + 100.0) * 100.0
    return (-o) / ((-o) + 100.0) * 100.0


SOFTMAX_TEMPERATURE_MULTIPLIER = 1.0  # T = multiplier * stdev(that week's L1 scores) — see backtest.py /
# temperature_analysis.py. A FIXED temperature was tested against 5 real 2026 tournaments (2 majors, 1
# marquee, 2 standard full-field events) and rejected: at a fixed T=8, the top pick's MDL PROB% averaged
# 17.16% and the top-5 combined averaged 45.69% across those events — implausibly high for fields of
# 67-132 scored players (no realistic PGA Tour favorite, even at their best, wins 15-20%+ of the time in a
# field that size). Rank correlation with actual finish (Spearman rho) is mathematically IDENTICAL at every
# temperature tested (8 through 25) — softmax is a strictly monotonic transform of L1, so temperature can
# only affect MDL PROB%'s magnitude/spread, never which player it favors. This is a pure calibration fix.
# An adaptive T = 1.0 x stdev(L1) landed at a mean top_prob of 5.72% (range 4.43-6.73%) and mean top-5
# cumulative of 20.66% across the same 5 events — a plausible range for real tournament win probabilities.
# Notable secondary finding: because L1 is percentile-based (0-100 rank transform), its week-to-week stdev
# was remarkably stable across 5 very different courses/field sizes (16.09-17.06) — so a well-chosen FIXED
# constant (~16) would likely have worked almost as well. The adaptive version is kept anyway since it's a
# small amount of extra code for a real guarantee rather than relying on a 5-sample regularity.


def softmax_probabilities(l1_scores: dict) -> dict:
    scored = {k: v for k, v in l1_scores.items() if v is not None}
    if not scored:
        return {}
    if len(scored) < 3:
        # Can't compute a meaningful stdev from fewer than 3 scores — fall
        # back to a fixed, moderate temperature rather than dividing by
        # something degenerate.
        temperature = 12.0
    else:
        temperature = SOFTMAX_TEMPERATURE_MULTIPLIER * statistics.pstdev(scored.values())
        if temperature <= 0:
            temperature = 12.0  # every player scored identically — degenerate spread, avoid division by zero
    top = max(scored.values())
    exps = {k: math.exp((v - top) / temperature) for k, v in scored.items()}
    total = sum(exps.values())
    return {k: e / total * 100.0 for k, e in exps.items()}


def compute_l3(l1_results: dict, market_data: dict) -> dict:
    win_resp = market_data["markets"]["outrights_win"]["response"]
    win_by_id = _index_by_dg_id(win_resp["odds"]) if win_resp and isinstance(win_resp.get("odds"), list) else {}

    probs = softmax_probabilities({k: v["l1_score"] for k, v in l1_results.items()})

    l3 = {}
    for dg_id, l1 in l1_results.items():
        if l1["l1_score"] is None:
            continue
        row = win_by_id.get(dg_id, {})
        book, odds_val = pick_book_odds(row)
        mkt_prob = implied_prob_pct(odds_val)
        mdl_prob = probs.get(dg_id)
        edge = (mdl_prob - mkt_prob) if (mdl_prob is not None and mkt_prob is not None) else None
        l3[dg_id] = {
            "l1_score": l1["l1_score"],
            "mdl_prob_pct": round(mdl_prob, 2) if mdl_prob is not None else None,
            "mkt_prob_pct": round(mkt_prob, 2) if mkt_prob is not None else None,
            "win_edge_pct": round(edge, 2) if edge is not None else None,
            "win_odds": odds_val,
            "win_book": book,
        }
    return l3


# ─────────────────────────────────────────────────────────────────────────
# Pick tiers — E/W Winner, Longshot/Value, Top 10/20, Matchup, Fade.
#
# FIRST-PASS HEURISTIC, NOT A PRECISELY SPECIFIED ALGORITHM. The CLAUDE.md
# design doc describes each tier's intent ("1-3 strong outright cases",
# "large model-rank-to-price gap", etc.) but not exact cutoffs — those are
# inherently judgment calls the real workbooks made case-by-case. This
# implementation picks reasonable, clearly-isolated thresholds so the
# output is inspectable and tunable, not a black box. Treat every constant
# in this section as provisional pending comparison against a real week's
# output alongside Brian's own judgment.
# ─────────────────────────────────────────────────────────────────────────
def assign_pick_tiers(l2_results: dict, l3: dict, market_data: dict) -> dict:
    passers = [dg_id for dg_id, r in l2_results.items() if r["pass"] and dg_id in l3]
    passers.sort(key=lambda d: l3[d]["l1_score"], reverse=True)

    tiers = {"ew_winner": [], "longshot_value": [], "top10_top20": [], "matchup": [], "fade": []}
    used = set()

    def odds_value(dg_id):
        v = l3[dg_id]["win_odds"]
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    # E/W Winner: top 3 L2-pass players by L1 rank with positive edge.
    for dg_id in passers:
        if len(tiers["ew_winner"]) >= 3:
            break
        edge = l3[dg_id]["win_edge_pct"]
        if edge is not None and edge > 0:
            tiers["ew_winner"].append(dg_id)
            used.add(dg_id)

    # Longshot / Value: L2-pass, not already used, long odds (>= +5000) with a real edge.
    longshots = [d for d in passers if d not in used and (odds_value(d) or 0) >= 5000
                 and (l3[d]["win_edge_pct"] or 0) > 0]
    longshots.sort(key=lambda d: l3[d]["win_edge_pct"], reverse=True)
    tiers["longshot_value"] = longshots[:2]
    used.update(tiers["longshot_value"])

    # Top 10 / Top 20: next-best L2-pass players by L1 rank, not already used.
    remaining = [d for d in passers if d not in used]
    tiers["top10_top20"] = remaining[:6]
    used.update(tiers["top10_top20"])

    # Matchup: only if matchups market is ready. Compares each match's two
    # players by L1 rank; flags it when the model's better-ranked player is
    # priced as the underdog in that specific matchup.
    if market_data.get("matchups_ready"):
        match_resp = market_data["markets"]["matchups"]["response"]
        seen_pairs = set()
        for m in (match_resp.get("match_list") or [])[:10]:
            p1, p2 = m.get("p1_dg_id"), m.get("p2_dg_id")
            if p1 not in l3 or p2 not in l3:
                continue
            pair_key = frozenset((p1, p2))
            if pair_key in seen_pairs:
                # Same two players can appear more than once in the raw feed
                # (e.g. separate round-matchup and tournament-matchup entries
                # for the same pairing) — one row per unique pair, not per entry.
                continue
            seen_pairs.add(pair_key)
            better, worse = (p1, p2) if l3[p1]["l1_score"] >= l3[p2]["l1_score"] else (p2, p1)
            better_odds, worse_odds = odds_value(better), odds_value(worse)
            if better_odds is not None and worse_odds is not None and better_odds > worse_odds:
                # better-ranked player priced as the underdog vs the worse-ranked one
                tiers["matchup"].append({"favor": better, "against": worse, "match": m})
            if len(tiers["matchup"]) >= 3:
                break

    # Fade: players priced as short favorites (<= +2000) with the most
    # negative edge — the market likes them more than the model does,
    # regardless of L2 status. Deliberately NOT gated on "L2 FAIL": a
    # fade is a market-price disagreement (model's implied win probability
    # is well below what the price demands), which an L2-PASS player can
    # trigger just as easily as an L2-FAIL one — real case found in the
    # 3M Open run: Scheffler at +270 with a -22.34% edge was L2 PASS and
    # got silently excluded by an earlier version of this condition that
    # required "not r['pass']", even though he's the textbook case the
    # tier exists for. L2 status still matters for the rationale text
    # (PASS players are "priced too short for the outright given the
    # model's actual probability", FAIL players are "market ignores what
    # the hard gates caught") but not for eligibility.
    fade_candidates = [d for d in l3
                        if (odds_value(d) or 99999) <= 2000
                        and (l3[d]["win_edge_pct"] or 0) < 0]
    fade_candidates.sort(key=lambda d: l3[d]["win_edge_pct"])
    tiers["fade"] = fade_candidates[:3]

    return tiers


# ─────────────────────────────────────────────────────────────────────────
# DFS — DraftKings salary join + lineup construction. Salaries come from
# dg_fantasy_projections() (see above), keyed by dg_id so they join
# directly onto the same metrics/l1_results/l2_results dicts everything
# else in this pipeline already uses — no name-string matching needed.
# ─────────────────────────────────────────────────────────────────────────
DFS_SALARY_CAP = 50_000   # DraftKings PGA classic slate: fixed $50K cap
DFS_ROSTER_SIZE = 6       # 6 golfers, no position requirements


def build_dfs_index(projections_resp: dict | None) -> dict:
    """dg_id -> {salary, proj_points_total, proj_ownership, value,
    site_name_id}. Empty dict if projections aren't released yet this
    firing — every caller must treat a missing dg_id as 'no DK salary
    data available', never invent or estimate one (see CLAUDE.md Data
    rules)."""
    if not projections_resp:
        return {}
    out = {}
    for p in projections_resp.get("projections", []):
        dg_id = p.get("dg_id")
        if dg_id is None:
            continue
        out[dg_id] = {
            "salary": p.get("salary"),
            "proj_points_total": p.get("proj_points_total"),
            "proj_ownership": p.get("proj_ownership"),
            "value": p.get("value"),
            "site_name_id": p.get("site_name_id"),
        }
    return out


def _knapsack_lineup(candidates: list, cap: int = DFS_SALARY_CAP, roster_size: int = DFS_ROSTER_SIZE) -> list:
    """candidates: list of (dg_id, salary, objective_value). Exact 0/1
    knapsack maximizing summed objective_value across exactly roster_size
    players with total salary <= cap. DK salaries are always round
    hundreds, so the DP is bucketed in $100 steps (~150 candidates x 6
    slots x 500 buckets — cheap) for an exact optimum every run rather
    than a greedy top-N that can silently bust the cap or leave real
    salary on the table."""
    step = 100
    budget = cap // step
    dp = [[(-1.0, []) for _ in range(budget + 1)] for _ in range(roster_size + 1)]
    dp[0][0] = (0.0, [])
    for dg_id, salary, obj in candidates:
        cost = salary // step
        if cost > budget:
            continue
        for count in range(roster_size - 1, -1, -1):
            for b in range(budget - cost, -1, -1):
                prev_val, prev_ids = dp[count][b]
                if prev_val < 0:
                    continue
                cand_val = prev_val + obj
                if cand_val > dp[count + 1][b + cost][0]:
                    dp[count + 1][b + cost] = (cand_val, prev_ids + [dg_id])
    best_val, best_ids = max(dp[roster_size], key=lambda x: x[0])
    return best_ids if best_val >= 0 else []


def build_dfs_lineups(metrics: dict, l1_results: dict, l2_results: dict, dfs_index: dict) -> dict:
    """Two lineups, matching the construction rules observed in past
    published StrokesEdge DFS pieces:
    - GPP: maximizes total L1 score (model conviction) across the FULL
      field, not gated by L2 PASS — the entire point of a leverage build
      is favoring players the model likes that DK's own market hasn't
      priced up yet, regardless of the outright Winner DNA gate.
    - Cash: maximizes total DK 'value' (points per $1,000 salary) among
      L2 PASS players only — floor/consistency over spike upside.
    Returns empty lists (never a fabricated lineup) when fewer than
    DFS_ROSTER_SIZE eligible candidates exist."""
    gpp_candidates = [(d, dfs_index[d]["salary"], l1_results[d]["l1_score"])
                       for d in metrics
                       if d in dfs_index and dfs_index[d]["salary"]
                       and d in l1_results and l1_results[d]["l1_score"] is not None]
    cash_candidates = [(d, dfs_index[d]["salary"], dfs_index[d]["value"])
                        for d in metrics
                        if d in dfs_index and dfs_index[d]["salary"] and dfs_index[d]["value"]
                        and l2_results.get(d, {}).get("pass")]
    gpp_ids = _knapsack_lineup(gpp_candidates) if len(gpp_candidates) >= DFS_ROSTER_SIZE else []
    cash_ids = _knapsack_lineup(cash_candidates) if len(cash_candidates) >= DFS_ROSTER_SIZE else []
    return {"gpp": gpp_ids, "cash": cash_ids}


# ─────────────────────────────────────────────────────────────────────────
# Excel workbook builder — 7 sheets, matching the format reference
# (StrokesEdge_OpenChampionship2026_MODEL_FULL_Customer Copy (1).xlsx).
# ─────────────────────────────────────────────────────────────────────────
FOOTER_STANDARD = ("Not financial advice. Gamble responsibly. | strokesedge.com/picks.html | "
                    "All picks logged to public tracker before first round.")
FOOTER_SHORT = "Not financial advice. Gamble responsibly. | strokesedge.com/picks.html"

TIER_LABELS = {
    "ew_winner": "E/W WINNER", "longshot_value": "LONGSHOT / VALUE",
    "top10_top20": "TOP 10 / TOP 20", "matchup": "MATCHUP", "fade": "FADE",
}

# ─────────────────────────────────────────────────────────────────────────
# Brand styling — every color/font below was read directly off real
# customer-copy workbooks (openpyxl cell.fill/cell.font inspection), not
# guessed or approximated from the hex codes alone. Source: Cover/Picks
# Card/Value Screen/Model Rankings sheets of
# StrokesEdge_ScottishOpen_2026_MODEL_Customer Copy (3).xlsx.
# ─────────────────────────────────────────────────────────────────────────
BRAND_FONT = "Calibri"
COLOR_BG_DARK = "FF080B07"          # title/subtitle/column-header background
COLOR_HEADER_GREEN = "FF6AB83A"     # title text, column header text
COLOR_SECTION_FILL = "FF1C3A14"     # section-label bar background (e.g. "TOP PICKS AT A GLANCE")
COLOR_DATA_FILL = "FFD5E8D4"        # data row background (light mint)
COLOR_DATA_TEXT = "FF1C3A14"        # default data row text (dark green)
COLOR_HIGHLIGHT_GREEN = "FF1C6B22"  # key numeric values within a data row (odds, edge%, weight%, PASS/PLAY)
COLOR_WHITE = "FFFFFFFF"            # subtitle text (tournament name line)
COLOR_MUTED_GREEN = "FFA8C4A0"      # meta line text (course/dates)
COLOR_FADE_RED = "FFC0392B"         # FADE verdict text specifically — confirmed real usage, not invented


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(color: str = None, size: float = 9, bold: bool = False) -> Font:
    return Font(name=BRAND_FONT, size=size, bold=bold, color=color)


def _style_row(ws, row_num: int, ncols: int, bg: str, font_color: str, bold: bool = False,
               size: float = 9, highlight_cols: set = None) -> None:
    """Applies fill + font to exactly the cells written in a row (not the
    sheet's full column range, which can be wider from other rows).
    highlight_cols: 0-indexed column offsets within the row that get
    COLOR_HIGHLIGHT_GREEN instead of font_color — the "key numbers" a
    reader's eye should land on (odds, edge%, weight%), matching the real
    workbooks' convention of highlighting numeric columns distinctly from
    descriptive text columns."""
    highlight_cols = highlight_cols or set()
    fill = _fill(bg)
    for i in range(1, ncols + 1):
        cell = ws.cell(row_num, i)
        cell.fill = fill
        c = COLOR_HIGHLIGHT_GREEN if (i - 1) in highlight_cols else font_color
        cell.font = _font(color=c, size=size, bold=bold)


def write_title_row(ws, text: str, size: float = 16) -> int:
    ws.append([text])
    _style_row(ws, ws.max_row, 1, COLOR_BG_DARK, COLOR_HEADER_GREEN, bold=True, size=size)
    return ws.max_row


def write_subtitle_row(ws, text: str, size: float = 13) -> int:
    ws.append([text])
    _style_row(ws, ws.max_row, 1, COLOR_BG_DARK, COLOR_WHITE, bold=True, size=size)
    return ws.max_row


def write_meta_row(ws, text: str, size: float = 10) -> int:
    ws.append([text])
    _style_row(ws, ws.max_row, 1, COLOR_BG_DARK, COLOR_MUTED_GREEN, bold=False, size=size)
    return ws.max_row


def write_section_row(ws, text: str, size: float = 11) -> int:
    ws.append([text])
    _style_row(ws, ws.max_row, 1, COLOR_SECTION_FILL, COLOR_HEADER_GREEN, bold=True, size=size)
    return ws.max_row


def write_header_row(ws, values: list, size: float = 8) -> int:
    ws.append(values)
    _style_row(ws, ws.max_row, len(values), COLOR_BG_DARK, COLOR_HEADER_GREEN, bold=True, size=size)
    return ws.max_row


def write_data_row(ws, values: list, highlight_cols: set = None, size: float = 9) -> int:
    ws.append(values)
    _style_row(ws, ws.max_row, len(values), COLOR_DATA_FILL, COLOR_DATA_TEXT, bold=True,
               size=size, highlight_cols=highlight_cols)
    return ws.max_row


def write_plain_row(ws, values: list) -> int:
    """No brand fill/font — used for blank spacer rows and footer text,
    matching how the real workbooks leave those unstyled."""
    ws.append(values)
    return ws.max_row


def set_cell_style(ws, row: int, col_1indexed: int, color: str, bold: bool = True, size: float = 9) -> None:
    """Post-write override for a single cell — used for the FADE verdict's
    red text, which sits inside an otherwise normally-styled data row."""
    ws.cell(row, col_1indexed).font = _font(color=color, size=size, bold=bold)


# ─────────────────────────────────────────────────────────────────────────
# Dashboard charts — matplotlib PNGs embedded into the sheet, per explicit
# instruction. Note: the real reference workbook (Scottish Open) actually
# used NATIVE Excel bar charts for these two, not embedded raster images —
# confirmed by inspecting ws._charts, and its "Model vs Market Edge" chart
# used a single solid fill color throughout (#1C3A14), not a green/red
# split by sign (checked for per-data-point color overrides — none exist).
# Building matplotlib images with a green/red split anyway, per the
# explicit instruction, not the literal reference file.
# ─────────────────────────────────────────────────────────────────────────
_HEX = lambda c: "#" + c[-6:]  # strip the leading alpha byte from an ARGB hex string for matplotlib


def _style_chart_axes(fig, ax) -> None:
    fig.patch.set_facecolor(_HEX(COLOR_BG_DARK))
    ax.set_facecolor(_HEX(COLOR_BG_DARK))
    ax.tick_params(colors=_HEX(COLOR_MUTED_GREEN), labelsize=8)
    for spine in ax.spines.values():
        spine.set_color(_HEX(COLOR_SECTION_FILL))
    ax.xaxis.label.set_color(_HEX(COLOR_MUTED_GREEN))
    ax.yaxis.label.set_color(_HEX(COLOR_MUTED_GREEN))


def generate_l1_chart(ranked: list, metrics: dict, l1_results: dict, out_path: Path, n: int = 15) -> Path:
    top = ranked[:n]
    names = [metrics[d]["player_name"] for d in reversed(top)]
    scores = [l1_results[d]["l1_score"] for d in reversed(top)]

    fig, ax = plt.subplots(figsize=(6.4, 4.2), dpi=150)
    _style_chart_axes(fig, ax)
    ax.barh(names, scores, color=_HEX(COLOR_HEADER_GREEN))
    ax.set_title(f"Top {len(top)} Model Rankings (L1 Score)", color=_HEX(COLOR_HEADER_GREEN),
                 fontsize=11, fontweight="bold", fontname=BRAND_FONT)
    ax.set_xlabel("L1 Score")
    # Deliberately NOT setting Calibri on the y-tick (player name) labels.
    # Confirmed by direct isolation while building this: on this system,
    # forcing Calibri onto tick-label Text objects (either via
    # label.set_fontname() or an rc_context) makes them render as blank —
    # data/color were correct, the glyphs just never made it into the
    # saved PNG. The chart TITLE with fontname=Calibri renders fine (a
    # different text-rendering code path), so it's kept there; tick
    # labels fall back to matplotlib's default sans-serif. This only
    # affects these two PNG images — every actual spreadsheet cell still
    # uses Calibri via openpyxl Font(), which has no such issue.
    fig.savefig(out_path, facecolor=fig.get_facecolor(), bbox_inches="tight")
    plt.close(fig)
    return out_path


def generate_edge_chart(l3: dict, l2_results: dict, metrics: dict, out_path: Path, n: int = 12) -> Path:
    candidates = [d for d, r in l2_results.items() if r["pass"] and d in l3
                  and l3[d]["win_edge_pct"] is not None]
    candidates.sort(key=lambda d: l3[d]["win_edge_pct"], reverse=True)
    top = candidates[:n]
    names = [metrics[d]["player_name"] for d in reversed(top)]
    edges = [l3[d]["win_edge_pct"] for d in reversed(top)]
    colors = [_HEX(COLOR_HEADER_GREEN) if e >= 0 else _HEX(COLOR_FADE_RED) for e in edges]

    fig, ax = plt.subplots(figsize=(6.4, 3.6), dpi=150)
    _style_chart_axes(fig, ax)
    ax.barh(names, edges, color=colors)
    ax.axvline(0, color=_HEX(COLOR_MUTED_GREEN), linewidth=0.8)
    ax.set_title("Model vs Market Edge — Top Value Plays (%)", color=_HEX(COLOR_HEADER_GREEN),
                 fontsize=11, fontweight="bold", fontname=BRAND_FONT)
    ax.set_xlabel("WIN EDGE %")
    # Deliberately NOT setting Calibri on the y-tick (player name) labels.
    # Confirmed by direct isolation while building this: on this system,
    # forcing Calibri onto tick-label Text objects (either via
    # label.set_fontname() or an rc_context) makes them render as blank —
    # data/color were correct, the glyphs just never made it into the
    # saved PNG. The chart TITLE with fontname=Calibri renders fine (a
    # different text-rendering code path), so it's kept there; tick
    # labels fall back to matplotlib's default sans-serif. This only
    # affects these two PNG images — every actual spreadsheet cell still
    # uses Calibri via openpyxl Font(), which has no such issue.
    fig.savefig(out_path, facecolor=fig.get_facecolor(), bbox_inches="tight")
    plt.close(fig)
    return out_path


def _odds_index(market_data: dict, key: str) -> dict:
    resp = market_data["markets"].get(key, {}).get("response")
    if resp and isinstance(resp.get("odds"), list):
        return _index_by_dg_id(resp["odds"])
    return {}


def _fmt_odds(odds_row: dict | None) -> str:
    if not odds_row:
        return "—"
    _, val = pick_book_odds(odds_row)
    if val is None:
        return "—"
    try:
        f = float(val)
        return f"{f:+.0f}"
    except (TypeError, ValueError):
        return str(val)


# Real bug found and fixed while reviewing live output: FACTOR_CATALOG spans
# wildly different natural scales — cf_approach_comp lives around 1e-4,
# cf_short_comp around 1e-2, the SG blends around 1e0, and the rate stats
# (bob_pct/gir_pct/scrambling_pct) around 1e1. A single fixed round(v, 3) /
# ":+.3f" rounds cf_approach_comp to exactly 0.000 for EVERY player — not a
# display quirk, the value is genuinely destroyed before it reaches the
# cell, even though the underlying percentile-based L1 scoring (which uses
# the unrounded value) was never affected. Confirmed live: 156 real,
# distinct cf_approach_comp values (range -0.000127 to +0.000203) all
# rounded to zero under the old fixed-3-decimals rule.
def _round_factor_value(v):
    """Rounds to enough decimal places to preserve ~3 significant figures
    regardless of magnitude, instead of a fixed decimal count. Returns a
    float (so the value stays numeric/sortable in Excel), or "N/A"."""
    if not isinstance(v, (int, float)):
        return "N/A"
    if v == 0:
        return 0.0
    magnitude = math.floor(math.log10(abs(v)))
    decimals = min(max(3, 2 - magnitude), 8)  # never fewer than 3, capped at 8 to avoid absurd precision
    return round(v, decimals)


def _fmt_factor_value(v) -> str:
    """String form of _round_factor_value, with an explicit sign — used in
    rationale text rather than a spreadsheet cell."""
    r = _round_factor_value(v)
    if r == "N/A":
        return "N/A"
    decimals = min(max(3, 2 - (math.floor(math.log10(abs(r))) if r != 0 else 0)), 8)
    return f"{r:+.{decimals}f}"


def _ordinal(n: float) -> str:
    """1st/2nd/3rd/4th... — found and fixed while reviewing real article
    output: the old code hardcoded 'th' for every value, producing '42th'
    and '23th' in actual published-quality text."""
    i = int(round(n))
    if 10 <= abs(i) % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(abs(i) % 10, "th")
    return f"{i}{suffix}"


def player_rationale(dg_id, metrics: dict, l1_results: dict, l2_results: dict, l3: dict, n: int = 3) -> str:
    """Deterministic, data-cited rationale — built from the actual scored
    components rather than a separate LLM call, so every pick's stated
    reason is directly traceable to a real number (per the 'every pick
    cites the specific stat values that support it' data rule), and so
    the pipeline isn't making a second unreviewed judgment call on top of
    the weight proposal."""
    comps = l1_results.get(dg_id, {}).get("components", {})
    top = sorted(comps.items(), key=lambda kv: kv[1]["weight_pct"], reverse=True)[:n]
    parts = [f"{FACTOR_CATALOG.get(k, k).split(' (')[0]} {_fmt_factor_value(v['value'])} ({_ordinal(v['percentile'])} pctile)"
             for k, v in top]
    l1 = l1_results.get(dg_id, {}).get("l1_score")
    l2 = l2_results.get(dg_id, {})
    edge = l3.get(dg_id, {}).get("win_edge_pct")
    bits = [f"L1 score {l1:.1f}" if l1 is not None else "L1 score unavailable"]
    bits.append("L2 PASS" if l2.get("pass") else "L2 FAIL")
    bits.extend(parts)
    if edge is not None:
        bits.append(f"WIN EDGE {edge:+.2f}%")
    return ". ".join(bits) + "."


def build_workbook(event: dict, ctx: dict) -> Path:
    metrics, l1_results, l2_results, l3 = ctx["metrics"], ctx["l1_results"], ctx["l2_results"], ctx["l3"]
    tiers, market_data = ctx["tiers"], ctx["market_data"]
    weights, gates = ctx["weights"], ctx["gates"]
    course_facts = ctx["course_facts"]
    year = date.fromisoformat(event["start_date"]).year

    win_idx = _odds_index(market_data, "outrights_win")
    top10_idx = _odds_index(market_data, "outrights_top_10")
    top20_idx = _odds_index(market_data, "outrights_top_20")

    wb = Workbook()
    tmp_chart_files = []

    # ---- 1. Cover ----------------------------------------------------
    ws = wb.active
    ws.title = "Cover"
    write_title_row(ws, "STROKESEDGE", 16)
    write_subtitle_row(ws, f"{event['event_name'].upper()} {year}  ·  MODEL OUTPUT", 13)
    meta_bits = [event.get("course_name") or "TBD", event.get("location") or "TBD",
                 event.get("start_date") or "TBD"]
    if course_facts.get("par"):
        meta_bits.append(f"Par {course_facts['par']}")
    if course_facts.get("yardage"):
        meta_bits.append(f"{course_facts['yardage']:,} Yds")
    write_meta_row(ws, "  ·  ".join(str(b) for b in meta_bits))
    write_plain_row(ws, [])
    write_section_row(ws, "TOP PICKS AT A GLANCE")
    write_header_row(ws, ["TIER", "PLAYER", "ODDS (FD)", "MDL RK", "EDGE", "KEY REASON"])

    ranked = sorted((d for d in l3 if l2_results.get(d, {}).get("pass")),
                     key=lambda d: l3[d]["l1_score"], reverse=True)
    rank_of = {d: i + 1 for i, d in enumerate(ranked)}

    for tier_key in ("ew_winner", "longshot_value", "top10_top20"):
        for dg_id in tiers[tier_key]:
            write_data_row(ws, [
                TIER_LABELS[tier_key], metrics[dg_id]["player_name"], _fmt_odds(win_idx.get(dg_id)),
                f"#{rank_of.get(dg_id, '—')}",
                f"{l3[dg_id]['win_edge_pct']:+.2f}%" if l3[dg_id]["win_edge_pct"] is not None else "—",
                player_rationale(dg_id, metrics, l1_results, l2_results, l3, n=2),
            ], highlight_cols={2, 4})

    # FADE and MATCHUPS are required sections in every published piece —
    # always render the header, even with zero qualifying players, rather
    # than silently omitting the section (found live: an in-progress
    # tournament's odds board can genuinely produce zero fade candidates —
    # real market behavior, not a bug — and silently dropping the section
    # made that indistinguishable from a broken tier).
    write_plain_row(ws, [])
    write_section_row(ws, "FADE")
    if tiers["fade"]:
        for dg_id in tiers["fade"]:
            row = write_data_row(ws, [
                "FADE (outright)", metrics[dg_id]["player_name"], _fmt_odds(win_idx.get(dg_id)),
                f"#{rank_of.get(dg_id, '—')}",
                f"{l3[dg_id]['win_edge_pct']:+.2f}%" if l3[dg_id]["win_edge_pct"] is not None else "—",
                "FADE OUTRIGHT ONLY — Top 10/20 still viable. " +
                player_rationale(dg_id, metrics, l1_results, l2_results, l3, n=2),
            ], highlight_cols={2, 4})
            set_cell_style(ws, row, 1, COLOR_FADE_RED)  # tier label reads "FADE (outright)" — flag it red
    else:
        write_plain_row(ws, ["No short favorites (odds <= +2000, negative edge) identified this week."])

    write_plain_row(ws, [])
    write_section_row(ws, "MATCHUPS")
    if tiers["matchup"]:
        for m in tiers["matchup"]:
            fav, opp = m["favor"], m["against"]
            write_data_row(ws, [
                "MATCHUP", f"{metrics[fav]['player_name']} vs {metrics[opp]['player_name']}",
                "—", "—", "—",
                f"Model favors {metrics[fav]['player_name']} (L1 {l1_results[fav]['l1_score']:.1f}) "
                f"over {metrics[opp]['player_name']} (L1 {l1_results[opp]['l1_score']:.1f}) "
                f"despite the market pricing it the other way.",
            ])
    elif not market_data.get("matchups_ready"):
        write_plain_row(ws, ["Matchup odds were never posted this week — no matchup markets to evaluate."])
    else:
        write_plain_row(ws, ["No matchups where the model disagreed with the market this week."])

    write_plain_row(ws, [])
    write_section_row(ws, "ABBREVIATIONS & GLOSSARY")
    for w in weights:
        write_plain_row(ws, [w["key"], FACTOR_CATALOG.get(w["key"], w["key"]), f"Weight {w['weight_pct']:g}%"])
    write_plain_row(ws, ["L1 SCORE", "Composite course-fit score, percentile-weighted blend of this week's factors (0-100)"])
    write_plain_row(ws, ["L2 GATE", "Winner DNA filter — PASS = clears every gate threshold for this course"])
    write_plain_row(ws, ["MDL PROB%", "Model-implied win probability (softmax over L1 scores)"])
    write_plain_row(ws, ["MKT PROB%", "Market-implied win probability (100/(odds+100) for + odds)"])
    write_plain_row(ws, ["WIN EDGE", "MDL PROB% minus MKT PROB% — positive means the model sees the player underpriced"])

    # ---- 2. Dashboard --------------------------------------------------
    ws = wb.create_sheet("Dashboard")
    write_title_row(ws, f"STROKESEDGE — {event['event_name'].upper()} {year} | DASHBOARD", 12)
    write_plain_row(ws, ["Visual summary of model output. Full detail lives in Model Rankings and Value Screen."])
    write_plain_row(ws, [])
    write_header_row(ws, ["PLAYER", "L1 SCORE"])
    for dg_id in ranked[:15]:
        write_data_row(ws, [metrics[dg_id]["player_name"], l1_results[dg_id]["l1_score"]], highlight_cols={1})

    chart_dir = event_dir(event["slug"])
    l1_chart_path = chart_dir / "_chart_l1_scores.png"
    edge_chart_path = chart_dir / "_chart_win_edge.png"
    try:
        generate_l1_chart(ranked, metrics, l1_results, l1_chart_path)
        ws.add_image(XLImage(str(l1_chart_path)), "E3")
        tmp_chart_files.append(l1_chart_path)
    except Exception as e:
        logger.error(f"[{event['slug']}] L1 chart generation failed, continuing without it: {e}")
    try:
        generate_edge_chart(l3, l2_results, metrics, edge_chart_path)
        ws.add_image(XLImage(str(edge_chart_path)), "E23")
        tmp_chart_files.append(edge_chart_path)
    except Exception as e:
        logger.error(f"[{event['slug']}] edge chart generation failed, continuing without it: {e}")

    # ---- 3. Picks Card ---------------------------------------------------
    ws = wb.create_sheet("Picks Card")
    write_title_row(ws, f"STROKESEDGE — {event['event_name'].upper()} {year} | PICKS CARD", 12)
    write_meta_row(ws, "  ·  ".join(str(b) for b in meta_bits))
    write_plain_row(ws, [])
    headers = ["TIER", "PLAYER", "BET TYPE", "ODDS (FD)", "STAKE", "MDL RK", "EDGE%", "VERDICT", "RATIONALE", "L2"]
    picks_card_header_row = write_header_row(ws, headers)
    bet_type = {"ew_winner": "Outright / E-W", "longshot_value": "Outright",
                "top10_top20": "Top 10/20 Finish", "matchup": "Tournament Matchup", "fade": "Outright Only"}
    for tier_key in ("ew_winner", "longshot_value", "top10_top20"):
        for dg_id in tiers[tier_key]:
            write_data_row(ws, [
                TIER_LABELS[tier_key], metrics[dg_id]["player_name"], bet_type[tier_key],
                _fmt_odds(win_idx.get(dg_id)), "—", f"#{rank_of.get(dg_id, '—')}",
                f"{l3[dg_id]['win_edge_pct']:+.2f}%" if l3[dg_id]["win_edge_pct"] is not None else "—",
                "PLAY", player_rationale(dg_id, metrics, l1_results, l2_results, l3),
                "PASS" if l2_results[dg_id]["pass"] else "FAIL",
            ], highlight_cols={3, 6, 7})
    for m in tiers["matchup"]:
        fav, opp = m["favor"], m["against"]
        fav_pass = l2_results.get(fav, {}).get("pass", False)
        opp_pass = l2_results.get(opp, {}).get("pass", False)
        both_pass = fav_pass and opp_pass
        # A matchup pick is only a genuine PLAY if both players clear L2 —
        # the favored side winning its Winner-DNA gates doesn't matter if the
        # underdog side is a known-bad-fit player the model wouldn't trust
        # outright either. Was hardcoded PLAY regardless of gate status (real
        # bug found 2026-07-28: a matchup could show PLAY while its own L2
        # column read FAIL). WATCH ONLY here, same convention as a single-
        # player L2 FAIL restricting them to finish markets only.
        verdict = "PLAY" if both_pass else "WATCH ONLY"
        rationale = (f"Model favors {metrics[fav]['player_name']} over {metrics[opp]['player_name']} "
                     f"against the market's own pricing of this matchup.")
        if not both_pass:
            failed_names = [metrics[d]["player_name"] for d, p in ((fav, fav_pass), (opp, opp_pass)) if not p]
            rationale += (f" WATCH ONLY — {' and '.join(failed_names)} "
                          f"fail{'s' if len(failed_names) == 1 else ''} this course's Winner DNA gate(s).")
        write_data_row(ws, [
            "MATCHUP", f"{metrics[fav]['player_name']} vs {metrics[opp]['player_name']}",
            bet_type["matchup"], "—", "—", "—", "—", verdict, rationale,
            "PASS" if both_pass else "FAIL",
        ], highlight_cols={7})
    for dg_id in tiers["fade"]:
        row = write_data_row(ws, [
            "FADE", metrics[dg_id]["player_name"], bet_type["fade"], _fmt_odds(win_idx.get(dg_id)),
            "—", f"#{rank_of.get(dg_id, '—')}",
            f"{l3[dg_id]['win_edge_pct']:+.2f}%" if l3[dg_id]["win_edge_pct"] is not None else "—",
            "FADE", "FADE OUTRIGHT ONLY — Top 10/20 still viable. " +
            player_rationale(dg_id, metrics, l1_results, l2_results, l3),
            # Was hardcoded "FAIL" — safe only while fade required L2 FAIL
            # by construction. Now a fade can be an L2-PASS player priced
            # too short for the outright (e.g. Scheffler, +270 / -22.34%
            # edge, L2 PASS) so this must reflect the player's actual gate
            # status, same as the MATCHUP row above.
            "PASS" if l2_results.get(dg_id, {}).get("pass") else "FAIL",
        ], highlight_cols={3, 6})  # VERDICT (col 8) deliberately excluded — styled red below, not highlight green
        set_cell_style(ws, row, 8, COLOR_FADE_RED)

    # Standing validation, not a one-time patch: VERDICT and L2 must never
    # disagree on a shipped row (added 2026-07-28 after the matchup-gate bug
    # above shipped PLAY on a row whose own L2 column read FAIL). Checked
    # generically across every data row in this sheet, not just matchups, so
    # any future tier added to Picks Card is covered by construction rather
    # than needing its own copy of this check.
    for r in range(picks_card_header_row + 1, ws.max_row + 1):
        verdict_val = ws.cell(row=r, column=8).value
        l2_val = ws.cell(row=r, column=10).value
        if verdict_val == "PLAY" and l2_val == "FAIL":
            player_val = ws.cell(row=r, column=2).value
            raise RuntimeError(
                f"[{event['slug']}] Picks Card row {r} ({player_val!r}) has VERDICT=PLAY but L2=FAIL — "
                f"refusing to ship an internally inconsistent workbook. Fix the tier logic that produced this row."
            )

    write_plain_row(ws, [])
    write_plain_row(ws, [FOOTER_STANDARD])

    # ---- 4. Value Screen ---------------------------------------------------
    ws = wb.create_sheet("Value Screen")
    write_title_row(ws, f"STROKESEDGE — {event['event_name'].upper()} {year} | L3 VALUE SCREEN", 12)
    gate_text = " · ".join(f"{g['key']} {g['operator']} {g['value']:g}" for g in gates)
    write_meta_row(ws, f"L2 PASS players sorted by WIN EDGE descending  ·  Gates: {gate_text}")
    write_plain_row(ws, [])
    factor_keys = [w["key"] for w in weights]
    headers = (["RK", "MDL RK", "PLAYER", "L1 SCORE", "MDL PROB%", "MKT PROB%", "WIN EDGE"]
               + [k.upper() for k in factor_keys] + ["WIN ODDS (FD)", "TOP 10 ODDS (FD)", "TOP 20 ODDS (FD)"])
    write_header_row(ws, headers)
    win_edge_col = 6  # 0-indexed offset of WIN EDGE
    odds_cols = {len(headers) - 3, len(headers) - 2, len(headers) - 1}  # last 3 columns
    near_pass = [d for d, r in l2_results.items() if not r["pass"] and d in l3
                 and rank_of.get(d, 999) <= 15 and d not in ranked]

    def _edge_sort_key(d):
        edge = l3[d]["win_edge_pct"]
        return (edge is None, -edge if edge is not None else 0.0)

    screen_rows = sorted(ranked + near_pass, key=_edge_sort_key)
    for i, dg_id in enumerate(screen_rows, 1):
        row = [i, f"#{rank_of.get(dg_id, '—')}", metrics[dg_id]["player_name"], l1_results[dg_id]["l1_score"],
               l3[dg_id]["mdl_prob_pct"], l3[dg_id]["mkt_prob_pct"], l3[dg_id]["win_edge_pct"]]
        for k in factor_keys:
            v = metrics[dg_id]["factors"].get(k)
            row.append(_round_factor_value(v))
        row += [_fmt_odds(win_idx.get(dg_id)), _fmt_odds(top10_idx.get(dg_id)), _fmt_odds(top20_idx.get(dg_id))]
        write_data_row(ws, row, highlight_cols={win_edge_col} | odds_cols)
    write_plain_row(ws, [])
    write_plain_row(ws, ["L2 PASS players sorted first; L2 FAIL rows below MDL RK 15 shown as near-pass, "
                          "finish markets only. WIN EDGE = model probability minus market probability."])
    write_plain_row(ws, [FOOTER_SHORT])

    # ---- 5. Model Rankings ---------------------------------------------------
    ws = wb.create_sheet("Model Rankings")
    write_title_row(ws, f"STROKESEDGE — {event['event_name'].upper()} {year} | FULL MODEL RANKINGS", 12)
    weight_str = " / ".join(f"{w['key']} {w['weight_pct']:g}%" for w in weights)
    l30_note = ctx.get("l30_window_note", "")
    write_meta_row(ws, f"Weights: {weight_str}  |  SG Source: Data Golf skill-ratings baseline blended 60/40 "
                       f"with computed L30 ({l30_note})  |  L2 PASS = clears all {len(gates)} Winner DNA gates")
    write_plain_row(ws, [])
    dfs_index = ctx.get("dfs_index", {})
    headers = (["RK", "PLAYER", "DK SALARY"] + [k.upper() for k in factor_keys]
               + ["L2 GATE", "MDL PROB%", "MKT PROB%", "WIN EDGE", "WIN ODDS (FD)", "TOP 10 ODDS (FD)", "TOP 20 ODDS (FD)"])
    write_header_row(ws, headers)
    dk_salary_col = 2  # 0-indexed offset of the DK SALARY column (RK=0, PLAYER=1, DK SALARY=2)
    l2_gate_col = 3 + len(factor_keys)  # 0-indexed offset of the L2 GATE column (shifted for DK SALARY)
    win_edge_col2 = l2_gate_col + 3
    odds_cols2 = {len(headers) - 3, len(headers) - 2, len(headers) - 1}
    # Deliberately NOT `ranked` (that list is L2-PASS-only, used elsewhere for
    # tier construction). This sheet is documented as "full field, one row per
    # player" — an L2 FAIL doesn't mean unscored, it means restricted to finish
    # markets (see L2 GATE column below), so it still belongs here. Using
    # `ranked` here silently dropped every L2 FAIL player from the sheet
    # entirely (real bug found 2026-07-28: Rocket Classic had 134 scored
    # players but only the 52 L2-PASS players ever reached this sheet).
    full_field_ranked = sorted(l1_results.keys(), key=lambda d: l1_results[d]["l1_score"], reverse=True)
    for i, dg_id in enumerate(full_field_ranked, 1):
        salary = dfs_index.get(dg_id, {}).get("salary")
        row = [i, metrics[dg_id]["player_name"], f"${salary:,}" if salary else "N/A"]
        for k in factor_keys:
            v = metrics[dg_id]["factors"].get(k)
            row.append(_round_factor_value(v))
        row += ["PASS" if l2_results[dg_id]["pass"] else "FAIL",
                l3[dg_id]["mdl_prob_pct"], l3[dg_id]["mkt_prob_pct"], l3[dg_id]["win_edge_pct"],
                _fmt_odds(win_idx.get(dg_id)), _fmt_odds(top10_idx.get(dg_id)), _fmt_odds(top20_idx.get(dg_id))]
        write_data_row(ws, row, highlight_cols={dk_salary_col, l2_gate_col, win_edge_col2} | odds_cols2)

    # ---- 6. Watch List ---------------------------------------------------
    ws = wb.create_sheet("Watch List")
    write_title_row(ws, f"STROKESEDGE — {event['event_name'].upper()} {year} | WATCH LIST (OUTSIDE REGRESSION)", 12)
    write_plain_row(ws, ["Players without a reliable Data Golf strokes-gained sample. Not scored in L1 regression."])
    write_plain_row(ws, [])
    write_header_row(ws, ["PLAYER", "WIN ODDS (FD)", "TOP 20 ODDS (FD)", "TOTAL SG", "NOTE"])
    for dg_id, m in metrics.items():
        if m["sample_ok"]:
            continue
        write_data_row(ws, [
            m["player_name"], _fmt_odds(win_idx.get(dg_id)), _fmt_odds(top20_idx.get(dg_id)), "N/A",
            "Insufficient Data Golf strokes-gained sample — outside L1/L2 regression",
        ], highlight_cols={1, 2})

    # ---- 7. Model Weights ---------------------------------------------------
    ws = wb.create_sheet("Model Weights")
    write_title_row(ws, f"STROKESEDGE — {(event.get('course_name') or event['event_name']).upper()} {year} | "
                        f"MODEL WEIGHTS, STAT SOURCES & L2 GATE REFERENCE", 11)
    supplement_note = "PGA Tour CSV supplement used this week" if ctx.get("pga_supplement_used") else "no PGA Tour CSV supplement this week — Data Golf only"
    dfs_note = "preds/fantasy-projection-defaults" if ctx.get("dfs_index") else "preds/fantasy-projection-defaults (not released yet this week)"
    write_meta_row(ws, f"Data Golf endpoints: preds/skill-ratings, preds/approach-skill, preds/player-decompositions, "
                       f"historical-raw-data/rounds, betting-tools/outrights, betting-tools/matchups, {dfs_note}. "
                       f"L30 window: {l30_note}. {supplement_note}.")
    write_plain_row(ws, [])
    write_header_row(ws, ["STAT/CATEGORY", "WEIGHT", "SOURCE", "RATIONALE", "L2 GATE", "GATE VALUE", "PRIORITY"])
    gate_by_key = {g["key"]: g for g in gates}
    for i, w in enumerate(weights, 1):
        g = gate_by_key.get(w["key"])
        write_data_row(ws, [
            FACTOR_CATALOG.get(w["key"], w["key"]), f"{w['weight_pct']:g}%", "Data Golf", w["rationale"],
            w["key"] if g else "—", f"{g['operator']} {g['value']:g}" if g else "Contextual", f"#{i}",
        ], highlight_cols={1})
    write_plain_row(ws, [])
    write_section_row(ws, f"L2 WINNER DNA GATES — {(event.get('course_name') or event['event_name']).upper()} {year}")
    for g in gates:
        write_plain_row(ws, [f"Gate — {g['key']}", f"{g['operator']} {g['value']:g}", g["rationale"]])
    write_plain_row(ws, [])
    write_plain_row(ws, [FOOTER_SHORT])

    for sheet in wb.worksheets:
        sheet.column_dimensions[get_column_letter(1)].width = 24
        sheet.column_dimensions[get_column_letter(2)].width = 28

    out_path = event_dir(event["slug"]) / f"StrokesEdge_{_pascal_slug(event['slug'])}_{year}_MODEL.xlsx"
    wb.save(out_path)
    for p in tmp_chart_files:
        p.unlink(missing_ok=True)
    logger.info(f"[{event['slug']}] workbook saved: {out_path}")
    return out_path


def _pascal_slug(slug: str) -> str:
    return "".join(part.capitalize() for part in slug.split("-"))


# ─────────────────────────────────────────────────────────────────────────
# Weather — Open-Meteo, not Data Golf (confirmed: Data Golf's API exposes
# no weather/wind/forecast data anywhere, checked directly against the
# docs). Open-Meteo needs no API key and no signup at all — genuinely free,
# global coverage (works for Royal Birkdale same as any US venue), and the
# lat/lon it needs are already present on every event from Data Golf's own
# get-schedule response. Zero new credentials to manage.
# ─────────────────────────────────────────────────────────────────────────
WEATHER_BASE = "https://api.open-meteo.com/v1/forecast"
KMH_TO_MPH = 0.621371
C_TO_F = lambda c: c * 9.0 / 5.0 + 32.0


def fetch_weather_summary(event: dict) -> dict | None:
    """Returns a compact summary for the tournament's actual playing days
    (start_date through start_date+3, the usual Thu-Sun window), or None
    if lat/lon is missing or the call fails — weather is a nice-to-have
    for the article, never a reason to block the workbook/article
    pipeline."""
    lat, lon = event.get("latitude"), event.get("longitude")
    if lat is None or lon is None:
        logger.info(f"[{event['slug']}] no lat/lon on this event — skipping weather")
        return None
    try:
        start = date.fromisoformat(event["start_date"])
        end = start + timedelta(days=3)
        days_out = (end - date.today()).days
        if days_out < 0 or days_out > 15:
            logger.info(f"[{event['slug']}] tournament week is outside Open-Meteo's 16-day forecast "
                        f"range ({days_out} days out) — skipping weather")
            return None
        params = {
            "latitude": lat, "longitude": lon, "timezone": "auto",
            "forecast_days": min(days_out + 4, 16),
            "daily": "temperature_2m_max,temperature_2m_min,precipitation_sum,"
                     "windspeed_10m_max,windgusts_10m_max,winddirection_10m_dominant",
        }
        url = f"{WEATHER_BASE}?{urllib.parse.urlencode(params)}"
        with urllib.request.urlopen(url, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        daily = data.get("daily", {})
        times = daily.get("time", [])
        window_idx = [i for i, t in enumerate(times) if start.isoformat() <= t <= end.isoformat()]
        if not window_idx:
            return None
        wind_mph = [daily["windspeed_10m_max"][i] * KMH_TO_MPH for i in window_idx]
        gust_mph = [daily["windgusts_10m_max"][i] * KMH_TO_MPH for i in window_idx]
        temp_hi_f = [C_TO_F(daily["temperature_2m_max"][i]) for i in window_idx]
        temp_lo_f = [C_TO_F(daily["temperature_2m_min"][i]) for i in window_idx]
        precip_days = sum(1 for i in window_idx if daily["precipitation_sum"][i] > 1.0)
        return {
            "wind_mph_range": (round(min(wind_mph)), round(max(wind_mph))),
            "gust_mph_max": round(max(gust_mph)),
            "temp_f_range": (round(min(temp_lo_f)), round(max(temp_hi_f))),
            "precip_days": precip_days,
            "days_covered": len(window_idx),
        }
    except Exception as e:
        logger.error(f"[{event['slug']}] weather fetch failed, continuing without it: {e}")
        return None


# ─────────────────────────────────────────────────────────────────────────
# Substack picks article — built AFTER the workbook, from the workbook's
# own data (same ctx dict build_workbook() uses). Same design discipline
# as player_rationale(): every number in the article comes from code, not
# from the LLM. Claude writes ONLY the narrative connective tissue (intro
# hook, weather-implication prose, course-history color, closing) — one
# call, given the real numbers as context so it can reference them
# accurately without being asked to reproduce them from scratch.
#
# Structure and section order matched directly against two real published
# StrokesEdge articles (2026 Open Championship picks + DFS pieces) — not
# guessed. Voice rules (no em dashes, no AI-sounding parallel structure,
# no tidy summary sentences, uneven rhythm, abrupt section transitions)
# come from strokesedge-site/CLAUDE.md and strokesedge_opus_prompt.txt.
# ─────────────────────────────────────────────────────────────────────────
SITE_LINKS = {
    "picks": "https://strokesedge.com/picks.html",
    "analysis": "https://strokesedge.com/analysis.html",
    "methodology": "https://strokesedge.com/methodology.html",
    "gumroad": "https://strokesedge.gumroad.com/l/buehoc",
    "bmac": "https://buymeacoffee.com/strokesedge/membership",
}


def render_model_framework_section(ctx: dict) -> str:
    weights, gates = ctx["weights"], ctx["gates"]
    lines = ["## 01 — Model Framework", ""]
    for w in sorted(weights, key=lambda w: w["weight_pct"], reverse=True):
        label = FACTOR_CATALOG.get(w["key"], w["key"]).split(" (")[0]
        lines.append(f"- **{label} — {w['weight_pct']:g}%.** {w['rationale']}")
    lines.append("")
    lines.append("**Winner DNA gates this week:**")
    for g in gates:
        label = FACTOR_CATALOG.get(g["key"], g["key"]).split(" (")[0]
        lines.append(f"- {label} {g['operator']} {g['value']:g}. {g['rationale']}")
    return "\n".join(lines)


def render_top_model_outputs_section(event: dict, ctx: dict, win_idx: dict, ranked: list, rank_of: dict,
                                      writeups: dict) -> str:
    metrics, l1_results, l3 = ctx["metrics"], ctx["l1_results"], ctx["l3"]
    lines = ["## 03 — Top Model Outputs", ""]
    for dg_id in ranked[:3]:
        name = metrics[dg_id]["player_name"]
        odds = _fmt_odds(win_idx.get(dg_id))
        l1 = l1_results[dg_id]["l1_score"]
        mdl = l3[dg_id]["mdl_prob_pct"]
        edge = l3[dg_id]["win_edge_pct"]
        lines.append(f"**{name} ({odds})**")
        lines.append(f"Model rank #{rank_of[dg_id]}, L1 score {l1:.1f}, win probability {mdl:.2f}%"
                     + (f", edge {edge:+.2f}%." if edge is not None else "."))
        lines.append(writeups[dg_id]["long"])
        lines.append("")
    return "\n".join(lines)


def render_faq_section(event: dict, year: int, course_facts: dict, extra_qas: list) -> str:
    """Standing SEO section for every weekly article (picks + DFS), added
    2026-07-28 as a templated requirement rather than something requested
    per-week. Fully code-templated, not an LLM call: the questions are fixed
    and every answer is assembled from data already computed elsewhere in
    this run (dates, course facts, weights, top pick), so it can't drift or
    time out the way a Claude call could. extra_qas lets each caller add its
    own 1-2 article-specific pairs on top of the two universal ones below,
    keeping the total at 3-5 pairs."""
    course_name = event.get("course_name") or "the host course"
    location = event.get("location") or ""
    dates = event.get("start_date") or "TBD"
    course_desc = course_name + (f", {location}" if location else "")
    if course_facts.get("par") and course_facts.get("yardage"):
        course_desc += f" — par {course_facts['par']}, {course_facts['yardage']:,} yards"
    qas = [
        (f"When is the {event['event_name']}?",
         f"The {year} {event['event_name']} runs {dates}."),
        (f"What course is the {event['event_name']} played at?",
         f"{course_desc}."),
    ] + extra_qas
    lines = ["## FAQ", ""]
    for q, a in qas:
        lines += [f"**{q}**", a, ""]
    return "\n".join(lines).rstrip()


def render_full_picks_card_section(event: dict, ctx: dict, win_idx: dict, rank_of: dict, writeups: dict) -> str:
    metrics, l1_results = ctx["metrics"], ctx["l1_results"]
    tiers, market_data = ctx["tiers"], ctx["market_data"]
    lines = ["## 04 — Full Picks Card", ""]
    for tier_key in ("ew_winner", "longshot_value", "top10_top20"):
        if not tiers[tier_key]:
            continue
        lines.append(f"**{TIER_LABELS[tier_key]}**")
        for dg_id in tiers[tier_key]:
            lines.append(f"- {metrics[dg_id]['player_name']} ({_fmt_odds(win_idx.get(dg_id))}) — "
                         f"model rank #{rank_of.get(dg_id, '—')}. {writeups[dg_id]['short']}")
        lines.append("")

    # FADE and MATCHUPS are required sections in every published piece —
    # always render the header, even with zero qualifying players, rather
    # than silently omitting the section (same fix applied to the
    # workbook — see build_workbook for why this matters: an in-progress
    # tournament's odds board can genuinely produce zero fade candidates,
    # real market behavior, and silent omission made that indistinguishable
    # from a broken tier).
    lines.append("**MATCHUPS**")
    if tiers["matchup"]:
        for m in tiers["matchup"]:
            fav, opp = m["favor"], m["against"]
            lines.append(f"- {metrics[fav]['player_name']} over {metrics[opp]['player_name']} — "
                         f"model favors {metrics[fav]['player_name']} (L1 {l1_results[fav]['l1_score']:.1f}) "
                         f"against {metrics[opp]['player_name']} (L1 {l1_results[opp]['l1_score']:.1f}) "
                         f"despite the market pricing it the other way.")
    elif not market_data.get("matchups_ready"):
        lines.append("- Matchup odds were never posted this week. No matchup markets to evaluate.")
    else:
        lines.append("- No matchups where the model disagreed with the market this week.")
    lines.append("")

    lines.append("**FADE**")
    if tiers["fade"]:
        for dg_id in tiers["fade"]:
            lines.append(f"- {metrics[dg_id]['player_name']} ({_fmt_odds(win_idx.get(dg_id))}) — "
                         f"FADE OUTRIGHT ONLY, Top 10/20 still viable. {writeups[dg_id]['long']}")
    else:
        lines.append("- No short favorites (odds <= +2000, negative edge) identified this week.")
    lines.append("")
    return "\n".join(lines)


ARTICLE_NARRATIVE_PROMPT = """You are the voice behind StrokesEdge, a quantitative PGA Tour golf betting brand publishing on Substack. Write ONLY the four short narrative blocks below for this week's picks article — nothing else, no other sections, no headers.

TOURNAMENT: {event_name}
COURSE: {course_name}
LOCATION: {location}
PAR: {par}
YARDAGE: {yardage}
WEATHER: {weather_text}
TOP MODEL PICK: {top_pick_name} ({top_pick_odds}), L1 score {top_pick_l1:.1f}, {top_pick_reason}
COURSE HISTORY FACTOR WEIGHT THIS WEEK: {course_history_weight}

VOICE RULES — non-negotiable:
- No em dashes anywhere in the prose.
- No AI-sounding parallel structure ("not just X, but Y" repeated patterns).
- No tidy summary sentence closing every paragraph.
- Uneven rhythm — mix long analytical sentences with short, punchy ones. At least one abrupt transition between blocks.
- Data-forward, analytical, dismissive of conventional narrative, never hyperbolic.
- If you reference specific past-tournament history at this course (years, past winners), only do so if you're genuinely confident it's accurate — this is general knowledge context, not verified pipeline data, so flag internally by keeping it plausible and general if unsure rather than inventing specifics.
- If you state this course's par or yardage anywhere in the prose, use ONLY the PAR/YARDAGE values given above verbatim — never a number recalled from training data. These are already shown in the article header, so it's fine to describe the course qualitatively (short, long, tight, etc.) without restating the number at all.

Respond with exactly this format, no extra commentary:

INTRO_HOOK: <2-3 sentences opening the article, tied to this week's course storyline — what makes this course/week distinct>

WEATHER_NARRATIVE: <2-4 sentences turning the weather data into a modeling implication — why it matters for who contends>

COURSE_HISTORY_NOTE: <3-5 sentences of course history color relevant to this week's storyline>

CLOSING: <2-3 sentences closing the piece, reiterating the top pick and the core thesis>
"""


def generate_article_narrative(event: dict, ctx: dict, weather: dict | None, ranked: list, win_idx: dict) -> dict:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY environment variable is not set")

    metrics, l1_results, l2_results, l3 = ctx["metrics"], ctx["l1_results"], ctx["l2_results"], ctx["l3"]
    top_id = ranked[0]
    top_reason = player_rationale(top_id, metrics, l1_results, l2_results, l3, n=2)
    course_hist_w = next((w["weight_pct"] for w in ctx["weights"] if w["key"] == "course_history_adjustment"), None)

    if weather:
        weather_text = (f"wind {weather['wind_mph_range'][0]}-{weather['wind_mph_range'][1]} mph "
                        f"(gusts to {weather['gust_mph_max']}), temps {weather['temp_f_range'][0]}-"
                        f"{weather['temp_f_range'][1]}F, {weather['precip_days']} day(s) with meaningful rain "
                        f"in the forecast window")
    else:
        weather_text = "forecast unavailable this run — write generically about typical conditions for this venue/season"

    course_facts = ctx.get("course_facts", {})
    prompt = ARTICLE_NARRATIVE_PROMPT.format(
        event_name=event["event_name"], course_name=event.get("course_name") or "unknown",
        location=event.get("location") or "unknown", weather_text=weather_text,
        par=course_facts.get("par") or "unknown", yardage=course_facts.get("yardage") or "unknown",
        top_pick_name=metrics[top_id]["player_name"], top_pick_odds=_fmt_odds(win_idx.get(top_id)),
        top_pick_l1=l1_results[top_id]["l1_score"], top_pick_reason=top_reason,
        course_history_weight=f"{course_hist_w:g}%" if course_hist_w is not None else "not weighted this week",
    )
    body = json.dumps({
        # Was 1200 — too tight for the same reason generate_pick_writeups()
        # had to be raised from 4000 to 12000: this model emits a 'thinking'
        # content block ahead of 'text', drawing from the SAME max_tokens
        # budget, and thinking-token consumption is non-deterministic per
        # call. Confirmed live: a real 3M Open run truncated mid-sentence
        # on CLOSING (the last of 5 fields parsed) — "...with a clean" and
        # nothing after — which the old empty-field-only check accepted
        # as valid since the field wasn't blank, just cut off. Raised to
        # 4000 for real headroom; the stop_reason check and _looks_complete
        # validation below (same pattern as generate_pick_writeups()) mean
        # a truncated response is now caught and retried rather than
        # silently shipped.
        "model": CLAUDE_MODEL_WEIGHTS, "max_tokens": 4000,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", data=body,
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        method="POST",
    )
    expected_keys = ("INTRO_HOOK", "WEATHER_NARRATIVE", "COURSE_HISTORY_NOTE", "CLOSING")
    prose_keys = expected_keys

    # Found in testing: an otherwise well-formed Claude response can still
    # occasionally drop or empty one field (non-deterministic LLM output,
    # not a parser bug — verified the same regex correctly extracts all 5
    # fields from a normal response). Rather than silently ship an article
    # with a blank section, validate every field is non-empty AND complete
    # (not truncated mid-sentence) and retry once before giving up — a real
    # failure here should be loud (caller already degrades gracefully to
    # workbook-only), never silent.
    last_parts = {}
    for attempt in range(2):
        with urllib.request.urlopen(req, timeout=90) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        if result.get("stop_reason") == "max_tokens":
            logger.error(f"Article narrative call hit max_tokens (attempt {attempt + 1}/2) — "
                         f"response was truncated, discarding and retrying")
            continue
        raw = "".join(block.get("text", "") for block in result.get("content", []))

        parts = {}
        for key in expected_keys:
            m = re.search(rf"{key}:\s*(.+?)(?=\n[A-Z_]+:|\Z)", raw, re.S)
            value = m.group(1).strip() if m else ""
            if value and key in prose_keys and not _looks_complete(value):
                value = ""  # cut off mid-sentence — treat as missing, not "found"
            parts[key.lower()] = value
        last_parts = parts

        missing = [k for k in expected_keys if not parts[k.lower()]]
        if not missing:
            return parts
        logger.error(f"Article narrative call returned empty/truncated field(s) {missing} "
                     f"(attempt {attempt + 1}/2){' — retrying' if attempt == 0 else ''}")

    raise RuntimeError(f"Article narrative missing required field(s) after retry: "
                       f"{[k for k in expected_keys if not last_parts[k.lower()]]}")


def render_weather_section(weather: dict | None) -> str:
    if not weather:
        return "## 02 — Weather\n\nForecast unavailable at build time."
    lo, hi = weather["wind_mph_range"]
    tlo, thi = weather["temp_f_range"]
    return (f"## 02 — Weather\n\n"
            f"Wind running {lo}-{hi} mph across the tournament window, gusts up to {weather['gust_mph_max']} mph. "
            f"Temperatures {tlo}-{thi}F. "
            f"{weather['precip_days']} day(s) in the window carrying meaningful rain.")


# ─────────────────────────────────────────────────────────────────────────
# Player write-up prose — real bug found in review: the article was using
# player_rationale() (built for the workbook's scannable audit-trail
# cells) for its prose sections too, producing the identical template
# ("model rank #X, edge +Y%. L1 score Z...") repeated verbatim per player.
# Fine for a spreadsheet cell, wrong for an article meant to read like
# analysis. This generates genuinely varied prose instead — same
# discipline as every other LLM call in this pipeline: Claude gets the
# real numbers as input and writes ABOUT them, it never invents a number.
# The field_best flag is computed here, not guessed by the model, so a
# superlative claim like "best in the field" is only made when literally
# true.
# ─────────────────────────────────────────────────────────────────────────
def _compute_field_bests(l1_results: dict) -> dict:
    bests = {}
    for dg_id, r in l1_results.items():
        for key, comp in r.get("components", {}).items():
            if comp.get("value") is None:
                continue
            if key not in bests or comp["value"] > l1_results[bests[key]]["components"][key]["value"]:
                bests[key] = dg_id
    return bests


def _pick_citation_data(dg_id, metrics: dict, l1_results: dict, l3: dict, field_bests: dict, n: int = 3) -> dict:
    comps = l1_results.get(dg_id, {}).get("components", {})
    top = sorted(comps.items(), key=lambda kv: kv[1]["weight_pct"], reverse=True)[:n]
    stats = [{
        "label": FACTOR_CATALOG.get(k, k).split(" (")[0],
        "value": _fmt_factor_value(v["value"]),
        "percentile": _ordinal(v["percentile"]),
        "field_best": field_bests.get(k) == dg_id,
    } for k, v in top]
    return {
        "name": metrics[dg_id]["player_name"],
        "l1_score": l1_results.get(dg_id, {}).get("l1_score"),
        "edge_pct": l3.get(dg_id, {}).get("win_edge_pct"),
        "stats": stats,
    }


def _looks_complete(text: str) -> bool:
    """A response cut off mid-generation ends mid-word/mid-clause with no
    terminal punctuation. Shared by every Claude call in this pipeline that
    parses free-text fields out of a response — this is the only reliable
    signal that distinguishes a truncated field (which a naive regex would
    happily accept as "found") from a genuinely complete one. Originally
    added for generate_pick_writeups() after a real truncated SHORT field
    parsed successfully under the old code; generate_article_narrative()
    had the identical gap (no stop_reason check, no completeness check) and
    hit the identical failure mode in production — Section 06 of a real
    3M Open article was cut off mid-sentence ("...with a clean") because
    the CLOSING field, parsed last, absorbed a truncated response with no
    detection. Both call sites now share this one check."""
    return bool(re.search(r'[.!?]["\')\]]?\s*$', text.strip()))


PICK_WRITEUP_PROMPT = """You are writing player-by-player betting analysis for a StrokesEdge Substack picks article. For EACH player listed below, write TWO versions of the analysis: a LONG version for a deep-dive section and a SHORT version for a compact scannable picks list. They must not be the same sentences reused or trimmed — write them as genuinely separate pieces of prose, since the same wording appearing twice in one article reads as templated.

RULES — strict, not stylistic suggestions:
- Cite ONLY the specific stat values given below for each player. Never invent a number, never invent a stat not listed for that player.
- Only claim a stat is "the best in the field" or any other superlative if that specific stat entry says FIELD BEST. Never claim a superlative that isn't marked true.
- Vary sentence structure and the opening word/phrase across players — no two players' LONG writeups should start the same way, and don't reuse the same sentence template player to player. Same rule applies separately among the SHORT writeups.
- No em dashes anywhere.
- No AI-sounding parallel structure ("not just X, but Y" repeated patterns).
- Uneven rhythm — mix a short punchy sentence with a longer analytical one, and vary which comes first player to player.
- Actually analyze what the numbers mean for this player's case, don't just list them in sentence form.
- LONG: 2-3 sentences, for a narrative deep-dive.
- SHORT: exactly 1 sentence, for a compact list entry — a genuinely different sentence from the LONG version, not a truncation of it.

PLAYERS:
{player_blocks}

Respond with exactly one block per player, this format, nothing else:
ID: <id number>
LONG: <2-3 sentences>
SHORT: <1 sentence, different wording from LONG>
"""


def generate_pick_writeups(player_data: dict) -> dict:
    """player_data: dg_id -> citation dict from _pick_citation_data(). One
    Claude call covering every player needed in the article at once —
    covers ALL of them in a single call specifically so the model can vary
    rhythm/openings against each other; separate per-player calls couldn't
    guarantee that. Returns dg_id -> {"long": ..., "short": ...} (may be
    missing entries if the call fails twice; caller fills any gap)."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY environment variable is not set")

    ids = list(player_data.keys())
    id_map = {i + 1: dg_id for i, dg_id in enumerate(ids)}
    blocks = []
    for i, dg_id in enumerate(ids, 1):
        d = player_data[dg_id]
        stat_lines = "; ".join(
            f"{s['label']} {s['value']} ({s['percentile']} percentile in field"
            + (", FIELD BEST" if s["field_best"] else "") + ")"
            for s in d["stats"]
        )
        edge_str = f"{d['edge_pct']:+.2f}%" if d["edge_pct"] is not None else "not available"
        l1_str = f"{d['l1_score']:.1f}" if d["l1_score"] is not None else "unavailable"
        blocks.append(f"ID {i}: {d['name']}. L1 score {l1_str}. Edge {edge_str}. Stats: {stat_lines}.")
    prompt = PICK_WRITEUP_PROMPT.format(player_blocks="\n".join(blocks))

    body = json.dumps({
        # 4000 was the original budget and looked plenty on paper (11
        # players x ~150-200 tokens of real LONG+SHORT text is well under
        # it), but this model returns a 'thinking' content block ahead of
        # the 'text' block, and thinking tokens draw from the SAME
        # max_tokens budget. Confirmed directly: one real call finished
        # cleanly using 1404 thinking + ~2000 text tokens for 11 players;
        # another call on the identical prompt burned far more into
        # thinking (non-deterministic) and cut the text off mid-word after
        # only ~4 players. Raised well past any thinking budget observed
        # so far, with real margin for a bigger field (more E/W Winner/
        # Top 10-20 entries some weeks) — a few extra cents per week is a
        # trivial cost next to silently losing most of the article.
        # Raised again from 8000: a real run still truncated mid-sentence
        # on the LAST player even at 8000 (thinking-token consumption is
        # non-deterministic per call, confirmed earlier), because the old
        # parser had no way to detect a truncated-but-present SHORT field
        # and silently accepted it. Fixed below with a stop_reason check
        # and a terminal-punctuation validation; 12000 gives real headroom
        # on top of that so the retry path is rarely needed.
        "model": CLAUDE_MODEL_WEIGHTS, "max_tokens": 12000,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", data=body,
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        method="POST",
    )

    writeups = {}
    for attempt in range(2):
        with urllib.request.urlopen(req, timeout=120) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        if result.get("stop_reason") == "max_tokens":
            logger.error(f"Pick writeup call hit max_tokens (attempt {attempt + 1}/2) — "
                         f"response was truncated, discarding and retrying")
            continue
        raw = "".join(block.get("text", "") for block in result.get("content", []))

        # Two-stage parsing, not one combined regex — verified directly
        # against a real captured 11-player response that the single-pass
        # ID/LONG/SHORT pattern matches perfectly when the format is clean
        # (confirmed: 11/11 matched), yet a real production call still
        # came back with most players missing. Root cause: occasional
        # Claude formatting drift on ONE player, combined with a single
        # combined regex, cascades — one malformed block can break
        # boundary detection for every player after it. Splitting into
        # (1) find "ID: N" boundaries, a simple/robust pattern, then (2)
        # extract LONG/SHORT independently within each isolated block
        # means a single malformed player degrades to just that player,
        # never cascades, and even a half-malformed block (e.g. missing
        # SHORT) still recovers by reusing whichever field parsed.
        for block_m in re.finditer(r"ID:\s*(\d+)\s*\n(.*?)(?=\nID:\s*\d+\s*\n|\Z)", raw, re.S | re.I):
            idx = int(block_m.group(1))
            if idx not in id_map or id_map[idx] in writeups:
                continue
            block_text = block_m.group(2)
            long_m = re.search(r"LONG:\s*(.+?)(?=\nSHORT:|\Z)", block_text, re.S | re.I)
            short_m = re.search(r"SHORT:\s*(.+?)\Z", block_text, re.S | re.I)
            long_text = long_m.group(1).strip() if long_m else None
            short_text = short_m.group(1).strip() if short_m else None
            # Reject a field that was "found" by the regex but cut off
            # mid-sentence (no terminal punctuation) — treat it the same
            # as missing so the retry/fallback path below actually kicks
            # in instead of silently shipping a broken sentence.
            if long_text and not _looks_complete(long_text):
                long_text = None
            if short_text and not _looks_complete(short_text):
                short_text = None
            if long_text and short_text:
                writeups[id_map[idx]] = {"long": long_text, "short": short_text}
            elif long_text or short_text:
                only = long_text or short_text
                writeups[id_map[idx]] = {"long": only, "short": only}

        missing = [dg_id for dg_id in ids if dg_id not in writeups]
        if not missing:
            return writeups
        logger.error(f"Pick writeup call missing {len(missing)}/{len(ids)} player(s) "
                     f"(attempt {attempt + 1}/2){' — retrying' if attempt == 0 else ''}")
    return writeups  # possibly incomplete — caller (generate_substack_article) fills any gap


# ─────────────────────────────────────────────────────────────────────────
# SEO metadata block — Brian's standing rule (2026-08-18, extended to six
# fields 2026-08-18 same day, see weekly-model/CLAUDE.md "StrokesEdge
# Article SEO Fields"): every picks article run must surface Article
# Title, SEO Title (<60 chars), SEO Description (50-160 chars, ends "Free
# analysis.", real odds), Email Subject, Tags, and Post URL Slug, labeled
# and ready to copy-paste into Substack. Same discipline as everything
# else in this pipeline — real odds and names pulled from l3/metrics,
# never invented, and a field that doesn't fit its length rule gets a
# loud log line instead of a silent truncation.
# ─────────────────────────────────────────────────────────────────────────
COURSE_NAME_SUFFIXES_TO_STRIP = (
    " Country Club", " Golf Club", " Golf Links", " Golf Course", " National Golf Club", " Links",
)

SLUG_STOPWORDS = {"the", "of", "a", "an"}


def _short_course_name(course_name: str) -> str:
    if not course_name:
        return course_name
    for suf in COURSE_NAME_SUFFIXES_TO_STRIP:
        if course_name.endswith(suf):
            return course_name[: -len(suf)].strip()
    return course_name


def _slugify(text: str) -> str:
    if not text:
        return ""
    words = [w for w in re.findall(r"[a-zA-Z0-9]+", text.lower()) if w not in SLUG_STOPWORDS]
    return "-".join(words)


def _player_first_last(player_name: str) -> str:
    # metrics[...]['player_name'] is stored "Last, First" — flip to
    # "First Last" for anything reader-facing (tags, email subject).
    if "," in player_name:
        last, first = (p.strip() for p in player_name.split(",", 1))
        return f"{first} {last}"
    return player_name


def _player_first_name(player_name: str) -> str:
    if "," in player_name:
        return player_name.split(",", 1)[1].strip()
    return player_name.split()[0] if player_name else player_name


def build_seo_fields(event: dict, year: int, article_title: str, l3: dict, l2_results: dict, metrics: dict) -> dict:
    course_name = event.get("course_name") or "TBD"
    short_course = _short_course_name(course_name)

    seo_title = f"{event['event_name']} {year} Picks, Best Bets & Odds | {course_name}"
    if len(seo_title) > 60:
        seo_title = f"{event['event_name']} {year} Picks, Best Bets & Odds | {short_course}"
    if len(seo_title) > 60:
        logger.error(f"[{event['slug']}] SEO title still {len(seo_title)} chars (>60) even with short course "
                     f"name {short_course!r} — trim manually before publishing: {seo_title!r}")

    def odds_val(dg_id):
        try:
            return float(l3[dg_id]["win_odds"])
        except (TypeError, ValueError, KeyError):
            return None

    # Value play: single biggest positive WIN EDGE among L2-PASS players —
    # the same real number the E/W Winner / Longshot tiers are built from.
    value_pool = [d for d, r in l2_results.items()
                  if r.get("pass") and d in l3 and (l3[d]["win_edge_pct"] or -999) > 0]
    value_id = max(value_pool, key=lambda d: l3[d]["win_edge_pct"]) if value_pool else None

    # Fade: same "short favorite, most negative edge" definition assign_pick_tiers() uses.
    fade_pool = [d for d in l3 if (odds_val(d) or 99999) <= 2000 and (l3[d]["win_edge_pct"] or 0) < 0]
    fade_id = min(fade_pool, key=lambda d: l3[d]["win_edge_pct"]) if fade_pool else None

    def _fmt_plain_odds(v):
        # l3[...]["win_odds"] is already a resolved American-odds number
        # (see compute_l3 -> pick_book_odds), not the odds_row dict
        # _fmt_odds() elsewhere expects — plain +/-NNN formatting here.
        try:
            return f"{float(v):+.0f}"
        except (TypeError, ValueError):
            return None

    value_odds = _fmt_plain_odds(odds_val(value_id)) if value_id is not None else None
    fade_odds = _fmt_plain_odds(odds_val(fade_id)) if fade_id is not None else None

    if value_odds and fade_odds:
        seo_description = (f"{year} {event['event_name']} picks, best bets, and odds at {course_name}. "
                            f"Quant model flags a {value_odds} value play and fades a {fade_odds} favorite. "
                            f"Free analysis.")
    else:
        # Real, if rare, case (thin field, no short favorites this week) —
        # never fabricate odds just to fill the template.
        logger.info(f"[{event['slug']}] SEO description: no qualifying value play and/or fade this week "
                     f"(value_id={value_id}, fade_id={fade_id}) — using odds-free fallback")
        seo_description = (f"{year} {event['event_name']} picks, best bets, and odds at {course_name}. "
                            f"Quant model course-fit rankings and full picks card. Free analysis.")

    if not (50 <= len(seo_description) <= 160):
        logger.error(f"[{event['slug']}] SEO description is {len(seo_description)} chars, outside the "
                     f"50-160 rule — trim/expand manually before publishing: {seo_description!r}")

    # Falls back to the top overall L1-ranked L2-pass player when no
    # positive-edge value pick exists this week — tags/email subject
    # should never go empty, and the model's top pick is a reasonable
    # substitute for "the value play" when nothing clears a positive edge.
    headline_id = value_id
    if headline_id is None:
        passers = [d for d, r in l2_results.items() if r.get("pass") and d in l3]
        headline_id = max(passers, key=lambda d: l3[d]["l1_score"]) if passers else None

    if headline_id is not None:
        headline_name_full = _player_first_last(metrics[headline_id]["player_name"])
        headline_first = _player_first_name(metrics[headline_id]["player_name"])
    else:
        headline_name_full = None
        headline_first = None

    # Email Subject — under 50 chars preferred (advisory, not a hard
    # length gate like SEO title/description above), must name something
    # specific (player + odds) rather than generic hype copy.
    if headline_first and value_odds:
        email_subject = f"{event['event_name']} picks are live: {headline_first} at {value_odds} is the play"
    elif headline_first:
        email_subject = f"{event['event_name']} picks are live: {headline_first} is the play"
    else:
        email_subject = f"{event['event_name']} picks are live: full card inside"
    if len(email_subject) > 50:
        logger.info(f"[{event['slug']}] Email subject is {len(email_subject)} chars (>50 preferred) — "
                     f"{email_subject!r}")

    # Tags — exactly 5, fixed order, last slot is the headline player's
    # full name (First Last), never left blank.
    tags = ["Golf", event["event_name"], "Golf Betting", "PGA Tour", headline_name_full or "TBD"]

    # Post URL Slug — event['slug'] is already Data Golf's own
    # lowercase-hyphenated tournament slug (e.g. "bmw-championship"), so
    # it's reused directly rather than re-derived, same "don't duplicate
    # a value that already exists" discipline as elsewhere in this file.
    post_url_slug = f"{event['slug']}-{year}-picks-best-bets-{_slugify(short_course)}"
    if len(post_url_slug) > 60:
        logger.info(f"[{event['slug']}] Post URL slug is {len(post_url_slug)} chars (>60) — {post_url_slug!r}")

    return {
        "article_title": article_title, "seo_title": seo_title, "seo_description": seo_description,
        "email_subject": email_subject, "tags": tags, "post_url_slug": post_url_slug,
    }


def render_seo_block(seo: dict) -> str:
    return ("=== SEO METADATA — copy into Substack ===\n"
            f"ARTICLE TITLE:   {seo['article_title']}\n"
            f"SEO TITLE:       {seo['seo_title']} ({len(seo['seo_title'])} chars)\n"
            f"SEO DESCRIPTION: {seo['seo_description']} ({len(seo['seo_description'])} chars)\n"
            f"EMAIL SUBJECT:   {seo['email_subject']} ({len(seo['email_subject'])} chars)\n"
            f"TAGS:            {', '.join(seo['tags'])}\n"
            f"POST URL:        {seo['post_url_slug']} ({len(seo['post_url_slug'])} chars)\n"
            "==========================================")


def generate_substack_article(event: dict, ctx: dict) -> Path:
    metrics, l1_results, l2_results, l3 = ctx["metrics"], ctx["l1_results"], ctx["l2_results"], ctx["l3"]
    market_data = ctx["market_data"]
    tiers = ctx["tiers"]
    win_idx = _odds_index(market_data, "outrights_win")
    ranked = sorted((d for d in l3 if l2_results.get(d, {}).get("pass")),
                     key=lambda d: l3[d]["l1_score"], reverse=True)
    rank_of = {d: i + 1 for i, d in enumerate(ranked)}

    # Every player who needs a prose write-up this run — top 3 plus every
    # picks-card tier entry (fade included, matchup excluded since those
    # are player-vs-player comparisons, not single-player citations).
    writeup_ids = list(dict.fromkeys(
        ranked[:3] + tiers["ew_winner"] + tiers["longshot_value"] + tiers["top10_top20"] + tiers["fade"]
    ))
    field_bests = _compute_field_bests(l1_results)
    citation_data = {dg_id: _pick_citation_data(dg_id, metrics, l1_results, l3, field_bests)
                      for dg_id in writeup_ids}
    try:
        writeups = generate_pick_writeups(citation_data)
    except Exception as e:
        logger.error(f"Pick writeup generation failed entirely, falling back to templated citations: {e}")
        writeups = {}
    missing = [dg_id for dg_id in writeup_ids if dg_id not in writeups]
    if missing:
        logger.error(f"[{event['slug']}] {len(missing)} player(s) falling back to templated rationale "
                     f"after writeup generation gap: {[metrics[d]['player_name'] for d in missing]}")
        for dg_id in missing:
            fallback = player_rationale(dg_id, metrics, l1_results, l2_results, l3, n=2)
            writeups[dg_id] = {"long": fallback, "short": fallback}

    weather = fetch_weather_summary(event)
    narrative = generate_article_narrative(event, ctx, weather, ranked, win_idx)

    year = date.fromisoformat(event["start_date"]).year
    course_facts = ctx.get("course_facts", {})
    meta_bits = [event.get("course_name") or "TBD", event.get("location") or "TBD", event.get("start_date") or "TBD"]
    if course_facts.get("par"):
        meta_bits.append(f"Par {course_facts['par']}")
    if course_facts.get("yardage"):
        meta_bits.append(f"{course_facts['yardage']:,} Yds")

    top_id = ranked[0]
    # Fixed title format (Brian, 2026-08-18) — proven to drive Google search
    # traffic, must not vary week to week. Code-templated, never left to a
    # Claude-written hook, same discipline as the lead sentence below.
    title = f"{event['event_name']} {year} Picks: Model Best Bets and Fades for {event.get('course_name') or 'TBD'}"

    seo = build_seo_fields(event, year, title, l3, l2_results, metrics)
    logger.info(f"[{event['slug']}] {render_seo_block(seo)}")

    # Guaranteed lead sentence, not left to the Claude-written intro_hook to
    # remember — standing SEO requirement added 2026-07-28: every weekly
    # article's first paragraph must name the tournament, year, and course,
    # every week, not just when flagged. Code-templated so it can't be
    # dropped by a prompt that drifts, same discipline as the par/yardage
    # fix above (verifiable facts come from code, not from hoping the model
    # states them).
    lead_sentence = (f"The {year} {event['event_name']} runs {event.get('start_date') or 'this week'} at "
                      f"{event.get('course_name') or 'the host course'}"
                      + (f", {event['location']}" if event.get("location") else "") + ".")

    top_weight = max(ctx["weights"], key=lambda w: w["weight_pct"])
    faq_extra = [
        (f"Who is StrokesEdge's top pick for the {event['event_name']}?",
         f"{metrics[top_id]['player_name']} ({_fmt_odds(win_idx.get(top_id))}). "
         + player_rationale(top_id, metrics, l1_results, l2_results, l3, n=1)),
        (f"How does the StrokesEdge model score players at {event.get('course_name') or 'this course'}?",
         f"The model weights {FACTOR_CATALOG.get(top_weight['key'], top_weight['key'])} highest this week at "
         f"{top_weight['weight_pct']:g}% — {top_weight['rationale']}"),
    ]

    sections = [
        render_seo_block(seo),
        "",
        f"# {title}",
        "  ·  ".join(str(b) for b in meta_bits),
        "",
        lead_sentence,
        "",
        narrative["intro_hook"],
        "",
        render_model_framework_section(ctx),
        "",
        render_weather_section(weather),
        "",
        narrative["weather_narrative"],
        "",
        render_top_model_outputs_section(event, ctx, win_idx, ranked, rank_of, writeups),
        render_full_picks_card_section(event, ctx, win_idx, rank_of, writeups),
        "## 05 — Course History Note",
        "",
        narrative["course_history_note"],
        "",
        "## 06 — The Number That Runs This Model",
        "",
        narrative["closing"],
        "",
        render_faq_section(event, year, course_facts, faq_extra),
        "",
        "---",
        f"Full betting record: {SITE_LINKS['picks']}  ·  Analysis: {SITE_LINKS['analysis']}  ·  "
        f"Methodology: {SITE_LINKS['methodology']}",
        f"Weekly model workbook: {SITE_LINKS['gumroad']} ($7)  ·  Membership: {SITE_LINKS['bmac']} ($21/mo)",
        "",
        "Not financial advice. Gamble responsibly.",
    ]
    article_md = "\n".join(sections)

    out_path = event_dir(event["slug"]) / f"substack_article_{_pascal_slug(event['slug'])}_{year}.md"
    out_path.write_text(article_md, encoding="utf-8")
    logger.info(f"[{event['slug']}] Substack article saved: {out_path}")
    return out_path


# ─────────────────────────────────────────────────────────────────────────
# DFS article — same design discipline as the picks article: every number
# (salaries, lineups, DK value, ownership, edge%) comes from code, Claude
# writes ONLY the narrative connective tissue. Structure and section order
# matched against a real published StrokesEdge DFS piece (Slate Setup /
# GPP Lineup / Cash Lineup / The Chalk Problem / Weather Note / Full
# Salary Board Reference) — see CLAUDE.md.
# ─────────────────────────────────────────────────────────────────────────
DFS_NARRATIVE_PROMPT = """You are the voice behind StrokesEdge, a quantitative PGA Tour golf betting brand publishing on Substack. Write ONLY the five short narrative blocks below for this week's DraftKings DFS article — nothing else, no other sections, no headers.

TOURNAMENT: {event_name}
COURSE: {course_name}
DRAFTKINGS SLATE: $50,000 cap, 6 golfers, no position requirements
GPP LINEUP: {gpp_summary}
CASH LINEUP: {cash_summary}
TOP CHALK FADE(S): {chalk_summary}
WEATHER: {weather_text}

VOICE RULES — non-negotiable:
- No em dashes anywhere in the prose.
- No AI-sounding parallel structure ("not just X, but Y" repeated patterns).
- No tidy summary sentence closing every paragraph.
- Uneven rhythm — mix long analytical sentences with short, punchy ones.
- Data-forward, analytical, dismissive of conventional narrative, never hyperbolic.
- Reference the specific players/numbers given above accurately — never invent a salary, value score, or ownership number not given here.

Respond with exactly this format, no extra commentary:

SLATE_SETUP: <2-4 sentences on this week's DK slate — cap structure, and what the model's weighting this week means for where salary efficiency shows up on the board>

GPP_RATIONALE: <3-5 sentences on the GPP lineup's construction logic, referencing the specific players/salaries/ownership given above>

CASH_RATIONALE: <2-4 sentences on the cash lineup's construction logic, referencing the specific players/values given above>

CHALK_NOTE: <2-4 sentences on the chalk fade(s) given above — why they're high-owned and why the model disagrees>

WEATHER_DFS_NOTE: <1-3 sentences turning the weather data into a DFS-specific implication — who it helps/hurts on this slate>
"""


def generate_dfs_narrative(event: dict, weather: dict | None, lineups: dict, dfs_index: dict,
                            metrics: dict, l3: dict) -> dict:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY environment variable is not set")

    def _fmt_val(v, fmt):
        return format(v, fmt) if v is not None else "—"

    def _lineup_summary(ids: list) -> str:
        if not ids:
            return "not enough eligible players this week to build one"
        parts = []
        for d in ids:
            row = dfs_index[d]
            parts.append(f"{metrics[d]['player_name']} (${row['salary']:,}, "
                         f"value {_fmt_val(row['value'], '.2f')}, "
                         f"{_fmt_val(row['proj_ownership'], '.1f')}% owned)")
        return "; ".join(parts)

    chalk = sorted((d for d in dfs_index if d in l3 and l3[d]["win_edge_pct"] is not None
                    and l3[d]["win_edge_pct"] < 0),
                   key=lambda d: dfs_index[d].get("proj_ownership") or 0, reverse=True)[:3]
    chalk_summary = ("; ".join(f"{metrics[d]['player_name']} "
                               f"({_fmt_val(dfs_index[d]['proj_ownership'], '.1f')}% owned, "
                               f"edge {l3[d]['win_edge_pct']:+.2f}%)" for d in chalk)
                     if chalk else "no high-owned players with a negative model edge this week")

    if weather:
        weather_text = (f"wind {weather['wind_mph_range'][0]}-{weather['wind_mph_range'][1]} mph "
                        f"(gusts to {weather['gust_mph_max']}), temps {weather['temp_f_range'][0]}-"
                        f"{weather['temp_f_range'][1]}F, {weather['precip_days']} day(s) with meaningful rain")
    else:
        weather_text = "forecast unavailable this run — write generically about typical conditions for this venue/season"

    prompt = DFS_NARRATIVE_PROMPT.format(
        event_name=event["event_name"], course_name=event.get("course_name") or "unknown",
        gpp_summary=_lineup_summary(lineups["gpp"]), cash_summary=_lineup_summary(lineups["cash"]),
        chalk_summary=chalk_summary, weather_text=weather_text,
    )
    body = json.dumps({
        "model": CLAUDE_MODEL_WEIGHTS, "max_tokens": 4000,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages", data=body,
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        method="POST",
    )
    expected_keys = ("SLATE_SETUP", "GPP_RATIONALE", "CASH_RATIONALE", "CHALK_NOTE", "WEATHER_DFS_NOTE")

    # Same retry/stop_reason/completeness-check pattern as
    # generate_article_narrative() — see that function's comments for why
    # a single max_tokens budget shared with the model's thinking block
    # makes this necessary rather than optional.
    last_parts = {}
    for attempt in range(2):
        with urllib.request.urlopen(req, timeout=90) as resp:
            result = json.loads(resp.read().decode("utf-8"))
        if result.get("stop_reason") == "max_tokens":
            logger.error(f"DFS narrative call hit max_tokens (attempt {attempt + 1}/2) — "
                         f"response was truncated, discarding and retrying")
            continue
        raw = "".join(block.get("text", "") for block in result.get("content", []))

        parts = {}
        for key in expected_keys:
            m = re.search(rf"{key}:\s*(.+?)(?=\n[A-Z_]+:|\Z)", raw, re.S)
            value = m.group(1).strip() if m else ""
            if value and not _looks_complete(value):
                value = ""  # cut off mid-sentence — treat as missing, not "found"
            parts[key.lower()] = value
        last_parts = parts

        missing = [k for k in expected_keys if not parts[k.lower()]]
        if not missing:
            return parts
        logger.error(f"DFS narrative call returned empty/truncated field(s) {missing} "
                     f"(attempt {attempt + 1}/2){' — retrying' if attempt == 0 else ''}")

    raise RuntimeError(f"DFS narrative missing required field(s) after retry: "
                       f"{[k for k in expected_keys if not last_parts[k.lower()]]}")


def render_dfs_lineup_table(ids: list, dfs_index: dict, rank_of: dict, metrics: dict) -> str:
    if not ids:
        return "*Not enough eligible players this week to build this lineup.*"
    lines = ["| Player | Salary | DK Value | Model Rank | Proj. Own% |", "|---|---|---|---|---|"]
    total_salary = 0
    for d in ids:
        row = dfs_index[d]
        total_salary += row["salary"]
        value_str = f"{row['value']:.2f}" if row["value"] is not None else "—"
        own_str = f"{row['proj_ownership']:.1f}%" if row["proj_ownership"] is not None else "—"
        lines.append(f"| {metrics[d]['player_name']} | ${row['salary']:,} | {value_str} | "
                     f"#{rank_of.get(d, '—')} | {own_str} |")
    remaining = DFS_SALARY_CAP - total_salary
    lines.append("")
    lines.append(f"**Total salary: ${total_salary:,} · ${remaining:,} remaining**")
    return "\n".join(lines)


def generate_dfs_article(event: dict, ctx: dict) -> Path | None:
    """Returns None (not an exception) when DK salaries simply aren't
    available yet this firing or the field is too thin for either
    lineup — a missing DFS article is an expected weekly outcome under
    those conditions, not a failure the caller needs to log as one."""
    metrics, l2_results, l3 = ctx["metrics"], ctx["l2_results"], ctx["l3"]
    l1_results = ctx["l1_results"]
    dfs_index = ctx.get("dfs_index", {})
    if not dfs_index:
        logger.info(f"[{event['slug']}] DK salaries not available this firing — skipping DFS article")
        return None

    lineups = build_dfs_lineups(metrics, l1_results, l2_results, dfs_index)
    if not lineups["gpp"] and not lineups["cash"]:
        logger.info(f"[{event['slug']}] not enough DK-priced players to build either DFS lineup — skipping DFS article")
        return None

    ranked = sorted((d for d in l3 if l2_results.get(d, {}).get("pass")),
                     key=lambda d: l3[d]["l1_score"], reverse=True)
    rank_of = {d: i + 1 for i, d in enumerate(ranked)}

    weather = fetch_weather_summary(event)
    narrative = generate_dfs_narrative(event, weather, lineups, dfs_index, metrics, l3)

    year = date.fromisoformat(event["start_date"]).year
    course_name = event.get("course_name") or "TBD"
    location = event.get("location") or "TBD"
    start_date = event.get("start_date") or "TBD"
    course_facts = ctx.get("course_facts", {})

    # Guaranteed lead sentence — same standing SEO requirement and same
    # code-templated (not Claude-written) approach as the picks article.
    lead_sentence = (f"The {year} {event['event_name']} runs {start_date} at {course_name}"
                      + (f", {location}" if location and location != "TBD" else "") + ".")
    faq_extra = [
        (f"What is the DraftKings salary cap for the {event['event_name']} main slate?",
         f"${DFS_SALARY_CAP:,}, {DFS_ROSTER_SIZE} golfers, no position requirements."),
        (f"How are the GPP and Cash lineups built for the {event['event_name']}?",
         "GPP maximizes total model L1 score across the full field regardless of L2 gate status, for tournament "
         "leverage. Cash maximizes DraftKings value (points per $1,000 salary) restricted to L2-PASS players only, "
         "for floor and consistency."),
    ]

    sections = [
        # Fixed title format (Brian, 2026-08-18) — proven to drive Google
        # search traffic, must not vary week to week.
        f"# {event['event_name']} {year} DFS Picks: Best DraftKings Lineups for {course_name}",
        f"### {course_name} · {location} · {start_date} · DraftKings Main Slate · ${DFS_SALARY_CAP:,} Salary Cap",
        "",
        lead_sentence,
        "",
        "## 01 — Slate Setup",
        "",
        narrative["slate_setup"],
        "",
        "## 02 — GPP Lineup",
        "",
        render_dfs_lineup_table(lineups["gpp"], dfs_index, rank_of, metrics),
        "",
        narrative["gpp_rationale"],
        "",
        "## 03 — Cash Lineup",
        "",
        render_dfs_lineup_table(lineups["cash"], dfs_index, rank_of, metrics),
        "",
        narrative["cash_rationale"],
        "",
        "## 04 — The Chalk Problem",
        "",
        narrative["chalk_note"],
        "",
        "## 05 — Weather Note for DFS Builders",
        "",
        narrative["weather_dfs_note"],
        "",
        "## 06 — Full Salary Board Reference",
        "",
        "For lineup-building outside these two builds, full DK salaries, projected points, value, and "
        "StrokesEdge model rank for every player on the slate are in this week's model workbook, "
        "the same Model Rankings sheet behind both lineups above.",
        "",
        render_faq_section(event, year, course_facts, faq_extra),
        "",
        "---",
        f"Full betting record: {SITE_LINKS['picks']}  ·  Analysis: {SITE_LINKS['analysis']}  ·  "
        f"Methodology: {SITE_LINKS['methodology']}",
        f"Weekly model workbook: {SITE_LINKS['gumroad']} ($7)  ·  Membership: {SITE_LINKS['bmac']} ($21/mo)",
        "",
        "Not financial advice. Gamble responsibly. DFS lineups reflect model output and salary "
        "efficiency, not guarantees of performance.",
    ]
    article_md = "\n".join(sections)

    out_path = event_dir(event["slug"]) / f"dfs_article_{_pascal_slug(event['slug'])}_{year}.md"
    out_path.write_text(article_md, encoding="utf-8")
    logger.info(f"[{event['slug']}] DFS article saved: {out_path}")
    return out_path


# ─────────────────────────────────────────────────────────────────────────
# Email — same SMTP/GMAIL_APP_PASSWORD pattern as weekly_course_update.py,
# extended to support an optional file attachment.
# ─────────────────────────────────────────────────────────────────────────
EMAIL_TO = "strokesedge@gmail.com"


def send_email(subject: str, body: str, attachment_path: Path | None = None,
               attachment_paths: list | None = None) -> None:
    """attachment_path: single-file convenience (existing call sites unchanged).
    attachment_paths: list of Paths, for when there's more than one (workbook + article)."""
    app_password = os.environ.get("GMAIL_APP_PASSWORD")
    if not app_password:
        logger.error("GMAIL_APP_PASSWORD not set — skipping email notification")
        return

    attachments = list(attachment_paths or [])
    if attachment_path:
        attachments.append(attachment_path)

    if attachments:
        msg = MIMEMultipart()
        msg.attach(MIMEText(body))
        for p in attachments:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(p.read_bytes())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{p.name}"')
            msg.attach(part)
    else:
        msg = MIMEText(body)

    msg["Subject"] = subject
    msg["From"] = EMAIL_TO
    msg["To"] = EMAIL_TO
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=30) as server:
            server.login(EMAIL_TO, app_password)
            server.sendmail(EMAIL_TO, [EMAIL_TO], msg.as_string())
        names = ", ".join(p.name for p in attachments)
        logger.info(f"Email sent: {subject}" + (f" (attached {names})" if attachments else ""))
    except Exception as e:
        logger.error(f"Failed to send email: {e}")


def render_proposal_email_body(event: dict, proposal: dict, file_path: Path) -> str:
    """The weights-approval email includes the full compact proposal
    directly in the body — Brian should be able to gut-check it without
    opening the file every time (explicit requirement, not just a
    courtesy)."""
    zero_warnings = detect_zeroed_core_factors(proposal)
    lines = [
        f"StrokesEdge weekly model — weight proposal for {event['event_name']} "
        f"({'main event' if event['is_main_event'] else 'opposite-field event'})",
    ]
    if zero_warnings:
        lines += ["", "!! ZEROED CORE FACTORS — CONFIRM THIS IS INTENTIONAL BEFORE APPROVING !!"]
        lines += [f"  - {w}" for w in zero_warnings]
    change_flag = proposal.get("course_change_flag", {})
    if change_flag.get("status") in ("yes", "uncertain"):
        lines += ["", "!! COURSE CHANGE FLAG — REVIEW COURSE HISTORY / COURSE EXPERIENCE WEIGHTS BEFORE APPROVING !!",
                  f"  - {change_flag['note']}"]
    lines += [
        "",
        "COURSE SUMMARY",
        proposal["summary"],
        "",
        "PROPOSED L1 WEIGHTS",
    ]
    for w in proposal["weights"]:
        lines.append(f"  {w['weight_pct']:>5.1f}%  {FACTOR_CATALOG.get(w['key'], w['key'])}")
        lines.append(f"          {w['rationale']}")
    lines += ["", "PROPOSED L2 GATES"]
    for g in proposal["gates"]:
        lines.append(f"  {g['key']} {g['operator']} {g['value']:g}  —  {g['rationale']}")
    lines += [
        "",
        f"File: {file_path}",
        "",
        "To approve as-is: open the file and change STATUS to APPROVED.",
        "To edit first: change any weight_pct or gate value directly on its line, then change STATUS to APPROVED.",
        "Leaving this unreviewed means no workbook for this event this week — the pipeline never auto-approves.",
    ]
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────
# Escalation — if a tournament is still stuck before the workbook is built
# by Wednesday evening (comfortably before Thursday tee times), send an
# explicit warning rather than silently retrying into the weekend.
# ─────────────────────────────────────────────────────────────────────────
def escalation_threshold(event: dict):
    """Wednesday 18:00 local time of tournament week — one day before a
    typical Thursday first round. Uses naive local time throughout, same
    convention as weekly_course_update.py (no timezone library dependency;
    Task Scheduler and this script both run in the machine's local time)."""
    start = date.fromisoformat(event["start_date"])
    d = start
    while d.weekday() != 2:  # walk back to the most recent Wednesday before start_date (Wednesday == 2)
        d -= timedelta(days=1)
    return datetime(d.year, d.month, d.day, 18, 0, 0)


def maybe_escalate(event: dict, state: dict) -> None:
    if state["step"] == "complete" or state.get("escalation_sent"):
        return
    if datetime.now() >= escalation_threshold(event):
        send_email(
            f"StrokesEdge: {event['event_name']} model may be delayed",
            f"'{event['event_name']}' is still stuck at pipeline step '{state['step']}' as of "
            f"{datetime.now().isoformat(timespec='minutes')}, past the Wednesday-evening checkpoint. "
            f"This tournament may not get a workbook this week if this doesn't clear soon.\n\n"
            f"State file: {state_path(event['slug'])}",
        )
        state["escalation_sent"] = True


# ─────────────────────────────────────────────────────────────────────────
# PGA Tour CSV supplement integration — SCOPED DOWN, READ BEFORE EXTENDING.
# The real column schema of Brian's actual PGA Tour CSV exports wasn't
# confirmed while building this (the one sample available had generic
# "custom stats N" headers, not confirmed real PGA Tour column names). So
# this v1 attaches the matched raw row for citation/audit purposes only —
# it is NOT wired into L1 weighting or FACTOR_CATALOG yet. Extending this
# once a real file's actual headers are known is an open follow-up, not
# done here to avoid guessing at a schema.
# ─────────────────────────────────────────────────────────────────────────
def attach_pga_supplement(metrics: dict, supplement: dict | None) -> None:
    if not supplement:
        return
    for dg_id, m in metrics.items():
        # Data Golf names are "Last, First"; PGA Tour CSV exports are
        # typically "First Last" — reconcile before matching.
        if "," in m["player_name"]:
            last, first = (p.strip() for p in m["player_name"].split(",", 1))
            reconciled = f"{first} {last}"
        else:
            reconciled = m["player_name"]
        row = supplement.get(reconciled) or supplement.get(m["player_name"])
        if row:
            m["pga_supplement_row"] = row


# ─────────────────────────────────────────────────────────────────────────
# State-machine steps
# ─────────────────────────────────────────────────────────────────────────
def _int_keyed(d: dict) -> dict:
    return {int(k): v for k, v in d.items()}


def _step_new(event: dict, state: dict, test_mode: bool) -> None:
    field = dg_field_updates(event["tour"])
    if not field or not field.get("field"):
        logger.info(f"[{event['slug']}] field not yet available (tour={event['tour']}) — will retry next firing")
        return

    course_facts = wikipedia_course_facts(event.get("course_name") or event["event_name"])
    pga_par, pga_yardage = pgatour_course_par_yardage(event["event_name"])
    if pga_par and pga_yardage:
        logger.info(f"[{event['slug']}] par/yardage sourced from pgatour.com course-stats "
                    f"(current-year, overriding Wikipedia par={course_facts.get('par')}/yardage={course_facts.get('yardage')})")
        course_facts["par"], course_facts["yardage"] = pga_par, pga_yardage
    is_major = is_major_event(event["event_name"])

    if test_mode:
        logger.info(f"[{event['slug']}] TEST MODE — field available, would call Claude API to propose "
                    f"weights now (course_facts={course_facts}, is_major={is_major}); nothing written.")
        return

    try:
        raw = propose_weights(event, course_facts, is_major)
        proposal = parse_claude_weight_response(raw)
    except Exception as e:
        logger.error(f"[{event['slug']}] weight proposal failed: {e}")
        return

    path = write_proposal(event, proposal)
    zero_warnings = detect_zeroed_core_factors(proposal)
    change_flag = proposal.get("course_change_flag", {})
    flagged = change_flag.get("status") in ("yes", "uncertain")
    subject_prefix = ("[COURSE CHANGE] " if flagged else "") + ("[ZEROED FACTOR] " if zero_warnings else "")
    send_email(f"{subject_prefix}StrokesEdge: {event['event_name']} — weight proposal ready for review",
               render_proposal_email_body(event, proposal, path))
    if zero_warnings:
        logger.info(f"[{event['slug']}] zeroed core factor(s) flagged: {'; '.join(zero_warnings)}")
    if flagged:
        logger.info(f"[{event['slug']}] course change flagged ({change_flag['status']}): {change_flag['note']}")
    state["step"] = "weights_proposed"
    state["course_facts"] = course_facts
    state["is_major"] = is_major


def _step_weights_proposed(event: dict, state: dict, test_mode: bool) -> bool:
    """Returns True if approval was found this firing (caller falls
    through to attempt L1/L2 in the same pass)."""
    approved = check_approval(event)
    if not approved:
        logger.info(f"[{event['slug']}] weights still pending review")
        return False
    if test_mode:
        logger.info(f"[{event['slug']}] TEST MODE — weights are approved, would proceed to L1/L2 now.")
        return False
    state["weights"] = approved["weights"]
    state["gates"] = approved["gates"]
    state["step"] = "weights_approved"
    logger.info(f"[{event['slug']}] weights approved — proceeding to L1/L2")
    return True


def _step_weights_approved(event: dict, state: dict, test_mode: bool) -> None:
    if test_mode:
        logger.info(f"[{event['slug']}] TEST MODE — would run L1/L2 regression now; nothing written.")
        return
    try:
        field = dg_field_updates(event["tour"])
        if not field or not field.get("field"):
            logger.info(f"[{event['slug']}] field pull failed at L1/L2 step — will retry next firing")
            return
        skill_ratings = dg_skill_ratings()
        approach_skill = dg_approach_skill()
        decompositions = dg_player_decompositions(event["tour"])
        as_of = date.today()
        l30_by_id, rate_by_id = compute_l30_window_stats(event["tour"], as_of)
        metrics = build_player_metrics(event, field, skill_ratings, approach_skill, decompositions,
                                        l30_by_id, rate_by_id)
        supplement = load_pga_supplement(event)
        attach_pga_supplement(metrics, supplement)

        l1_results = run_l1(metrics, state["weights"])
        l2_results = apply_l2(metrics, state["gates"])
    except Exception as e:
        logger.error(f"[{event['slug']}] L1/L2 regression failed: {e}")
        return

    state["metrics"] = metrics
    state["l1_results"] = l1_results
    state["l2_results"] = l2_results
    state["l30_window_note"] = f"{(as_of - timedelta(days=L30_WINDOW_DAYS)).isoformat()} to {as_of.isoformat()}"
    state["pga_supplement_used"] = supplement is not None
    state["step"] = "l1l2_complete"
    n_pass = sum(1 for r in l2_results.values() if r["pass"])
    logger.info(f"[{event['slug']}] L1/L2 complete — {len(l2_results)} scored, {n_pass} pass L2")


def _build_and_deliver(event: dict, state: dict, market_data: dict) -> None:
    """Builds the workbook + picks article + DFS article and emails all of
    them together. Shared by the normal step machine (_step_l1l2_complete,
    below) and --rebuild (main()), so every path that produces a
    deliverable for an event produces the same set of deliverables — a
    weekly firing and a manual rebuild can never drift apart."""
    metrics = _int_keyed(state["metrics"])
    l1_results = _int_keyed(state["l1_results"])
    l2_results = _int_keyed(state["l2_results"])

    l3 = compute_l3(l1_results, market_data)
    tiers = assign_pick_tiers(l2_results, l3, market_data)

    # DFS salaries — best-effort, separate from the odds-readiness gate
    # above. DK sometimes finalizes salaries a bit later in the week than
    # sportsbook odds/matchups do; a miss here degrades the DK SALARY
    # column to "N/A" and skips the DFS article for this firing, but must
    # never block the workbook/picks article, which are the primary
    # deliverable.
    dfs_projections = dg_fantasy_projections(event["tour"])
    dfs_index = build_dfs_index(dfs_projections)
    if dfs_index:
        logger.info(f"[{event['slug']}] DK salaries available for {len(dfs_index)} players this firing")
    else:
        logger.info(f"[{event['slug']}] DK salaries not available this firing "
                     f"(preds/fantasy-projection-defaults empty or not released yet)")

    ctx = {
        "metrics": metrics, "l1_results": l1_results, "l2_results": l2_results, "l3": l3,
        "tiers": tiers, "market_data": market_data,
        "weights": state["weights"], "gates": state["gates"],
        "course_facts": state.get("course_facts", {}),
        "l30_window_note": state.get("l30_window_note", "unknown"),
        "pga_supplement_used": state.get("pga_supplement_used", False),
        "dfs_index": dfs_index,
    }
    try:
        workbook_path = build_workbook(event, ctx)
    except Exception as e:
        logger.error(f"[{event['slug']}] workbook build failed: {e}")
        return

    passers = [d for d, r in l2_results.items() if r["pass"] and d in l1_results and l1_results[d]["l1_score"] is not None]
    top_pick = metrics[max(passers, key=lambda d: l1_results[d]["l1_score"])]["player_name"] if passers else "none"
    n_pass = len(passers)

    # Substack article — built from the same ctx/workbook data. A failure
    # here (e.g. a Claude API hiccup on the narrative call) must not lose
    # the workbook, which is the primary deliverable — log and continue
    # with just the workbook attached rather than blocking completion.
    attachments = [workbook_path]
    article_note = ""
    try:
        article_path = generate_substack_article(event, ctx)
        attachments.append(article_path)
        article_note = "\nSubstack article draft is attached too — review before publishing, same as always.\n"
    except Exception as e:
        logger.error(f"[{event['slug']}] Substack article generation failed, sending workbook only: {e}")
        article_note = "\n(Substack article generation failed this run — see log. Workbook is unaffected.)\n"

    # DFS article — same non-blocking discipline as the Substack article.
    # generate_dfs_article() itself returns None (not an exception) for
    # the expected "salaries not out yet" / "field too thin" cases, so
    # only genuine failures land in the except branch below.
    dfs_note = ""
    try:
        dfs_article_path = generate_dfs_article(event, ctx)
        if dfs_article_path:
            attachments.append(dfs_article_path)
            dfs_note = "\nDFS article draft is attached too — DraftKings GPP/cash lineups from this week's salaries.\n"
        else:
            dfs_note = "\n(No DFS article this run — DK salaries weren't available yet or the field was too thin. Workbook/picks article unaffected.)\n"
    except Exception as e:
        logger.error(f"[{event['slug']}] DFS article generation failed, continuing without it: {e}")
        dfs_note = "\n(DFS article generation failed this run — see log. Workbook/picks article unaffected.)\n"

    send_email(
        f"StrokesEdge: {event['event_name']} model workbook ready",
        f"The {event['event_name']} weekly model workbook is attached.\n\n"
        f"{len(l2_results)} players scored, {n_pass} passed L2.\n"
        f"Top model pick: {top_pick}\n"
        f"Matchups tier: {'included' if market_data['matchups_ready'] else 'skipped — matchup odds never posted this week'}\n"
        f"{article_note}"
        f"{dfs_note}",
        attachment_paths=attachments,
    )
    state["step"] = "complete"
    state["workbook_path"] = str(workbook_path)
    logger.info(f"[{event['slug']}] COMPLETE — workbook emailed")


def _step_l1l2_complete(event: dict, state: dict, test_mode: bool) -> None:
    market_data = check_markets(event["tour"])
    if not market_data["outrights_ready"]:
        logger.info(f"[{event['slug']}] odds not yet live for all consumed outright markets — retry next firing")
        return
    if test_mode:
        logger.info(f"[{event['slug']}] TEST MODE — odds are ready, would build the workbook now.")
        return
    _build_and_deliver(event, state, market_data)


def process_event(event: dict, test_mode: bool = False) -> None:
    state = load_state(event)
    step_at_start = state["step"]
    logger.info(f"[{event['slug']}] current step: {step_at_start}")

    if step_at_start == "new":
        _step_new(event, state, test_mode)
    elif step_at_start == "weights_proposed":
        if _step_weights_proposed(event, state, test_mode):
            _step_weights_approved(event, state, test_mode)
    elif step_at_start == "weights_approved":
        _step_weights_approved(event, state, test_mode)
    elif step_at_start == "l1l2_complete":
        _step_l1l2_complete(event, state, test_mode)
    elif step_at_start == "complete":
        logger.info(f"[{event['slug']}] already complete — nothing to do")
        return
    else:
        logger.error(f"[{event['slug']}] unknown state step '{step_at_start}' — leaving untouched, needs manual review")
        return

    if not test_mode:
        # Escalation is meant to catch a tournament that's genuinely stuck
        # across firings, not to fire in the same breath as real progress.
        # escalation_threshold() only knows the tournament's calendar (is
        # it past Wednesday evening?) — it has no way to see that this
        # exact firing just wrote a fresh proposal or just cleared L1/L2,
        # so it must be gated here, not inside maybe_escalate() itself.
        if state["step"] != step_at_start:
            logger.info(f"[{event['slug']}] progressed {step_at_start} -> {state['step']} this firing "
                        f"— skipping escalation check")
        else:
            maybe_escalate(event, state)
        save_state(event, state)


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--test", action="store_true",
                        help="Detect events and log what would happen at each step, writing/emailing nothing.")
    parser.add_argument("--event", metavar="SLUG",
                        help="Process only this tournament slug this firing (debugging a single event).")
    parser.add_argument("--rebuild", metavar="SLUG",
                        help="Force-rebuild and re-email an already-complete event's workbook + picks "
                             "article + DFS article from its saved state.json, bypassing the step "
                             "machine entirely (ignores state['step'], including 'complete'). For "
                             "backfilling a deliverable — e.g. a DFS article that didn't exist yet when "
                             "the event first completed — into an already-finished week. Requires "
                             "outright odds still be live for that event (aborts otherwise) and requires "
                             "state.json to already have metrics/l1_results/l2_results/weights/gates "
                             "(i.e. L1/L2 already ran at some point for this event).")
    args = parser.parse_args()

    try:
        acquire_lock()
    except AlreadyRunningError as e:
        logger.info(f"Exiting without running — {e}")
        return

    try:
        logger.info("=" * 70)
        logger.info(f"Run started {datetime.now().isoformat()} (test_mode={args.test})")
        try:
            events = detect_events_this_week()
        except Exception as e:
            logger.error(f"FATAL: could not detect this week's events: {e}")
            if not args.test:
                send_email("StrokesEdge: Weekly model pipeline FAILED — event detection failed",
                           f"Could not determine this week's PGA Tour event(s). No state was touched.\n\nError: {e}")
            return

        if args.rebuild:
            event = next((e for e in events if e["slug"] == args.rebuild), None)
            if not event:
                logger.error(f"--rebuild {args.rebuild!r} did not match any event Data Golf currently "
                             f"lists as upcoming — this only works while the event is still findable "
                             f"via detect_events_this_week() (i.e. before/during tournament week).")
                return
            state = load_state(event)
            if "metrics" not in state:
                logger.error(f"[{event['slug']}] --rebuild aborted — state.json has no metrics/l1_results "
                             f"yet (L1/L2 never ran for this event). Nothing to rebuild from.")
                return
            logger.info(f"[{event['slug']}] --rebuild requested — forcing workbook/article rebuild "
                        f"regardless of saved step (was '{state['step']}')")
            market_data = check_markets(event["tour"])
            if not market_data["outrights_ready"]:
                logger.error(f"[{event['slug']}] --rebuild aborted — outright odds are not currently "
                             f"live for this event, so odds/WIN EDGE columns can't be recomputed.")
                return
            _build_and_deliver(event, state, market_data)
            save_state(event, state)
            return

        if args.event:
            events = [e for e in events if e["slug"] == args.event]
            if not events:
                logger.error(f"--event {args.event!r} did not match any detected event this week")
                return

        for event in events:
            try:
                process_event(event, test_mode=args.test)
            except Exception as e:
                logger.error(f"[{event['slug']}] unhandled exception this firing: {e}")
                # One event's failure must not block the others — continue the loop.
                continue
    finally:
        release_lock()


if __name__ == "__main__":
    main()


# ═══════════════════════════════════════════════════════════════════════
# SETUP INSTRUCTIONS
# ═══════════════════════════════════════════════════════════════════════
#
# 1. Credentials — already set in this environment, nothing new to add:
#    DATAGOLF_API_KEY, ANTHROPIC_API_KEY, GMAIL_APP_PASSWORD are all
#    already Windows user environment variables (confirmed while building
#    this — weekly_course_update.py already depends on the latter two).
#    If setting up on a different machine:
#        [System.Environment]::SetEnvironmentVariable('DATAGOLF_API_KEY', '...', 'User')
#        [System.Environment]::SetEnvironmentVariable('ANTHROPIC_API_KEY', 'sk-ant-...', 'User')
#        [System.Environment]::SetEnvironmentVariable('GMAIL_APP_PASSWORD', '...', 'User')
#    Restart any open terminal / Task Scheduler won't see new values until you do.
#
# 2. Dependencies
#    ------------
#    Only third-party package needed is openpyxl (everything else is
#    stdlib, matching weekly_course_update.py's convention):
#        pip install openpyxl
#
# 3. Windows Task Scheduler — one task, repeating
#    -----------------------------------------------
#    a. Task Scheduler > Create Task (not "Basic Task" — need "Run whether
#       user is logged on or not").
#    b. General tab: name it "StrokesEdge Weekly Model Pipeline". Check
#       "Run whether user is logged on or not".
#    c. Triggers tab > New: Weekly, Monday, 6:00:00 AM. Then open
#       "Advanced settings" on that same trigger and check "Repeat task
#       every: 3 hours" with "for a duration of: 4 days". This one trigger
#       covers both field-availability polling (starting Monday morning)
#       and odds-availability polling (through Friday) — see CLAUDE.md
#       "Automation & Scheduling" for why this replaced two separate
#       schedules.
#    d. Actions tab > New:
#         Program/script:  C:\path\to\python.exe
#         Arguments:       weekly_model_pipeline.py
#         Start in:        C:\Users\bkopp\strokesedge-site\weekly-model
#       (Use the full path to python.exe from `where python` — Task
#       Scheduler does not use your shell's PATH.)
#    e. Conditions tab: uncheck "Start the task only if the computer is on
#       AC power" if this runs on a laptop.
#    f. Save. Test immediately via Task Scheduler > right-click the task >
#       Run, then check weekly_model_pipeline.log.
#
# 4. Per-tournament files, created automatically under weekly-model/[slug]/
#    -----------------------------------------------------------------------
#    state.json              — pipeline progress for that event
#    weights_proposal.md     — human-editable, STATUS line controls approval
#    pga_tour_supplement.csv — OPTIONAL, drop in manually any time before
#                               the L1/L2 step runs if you want the PGA
#                               Tour supplement that week (see CLAUDE.md)
#    StrokesEdge_[Tournament]_[Year]_MODEL.xlsx — final workbook, once built
#
# 5. Test mode
#    ---------
#        python weekly_model_pipeline.py --test
#    Detects this week's event(s) and logs what each one would do at its
#    current step, without calling the Claude API, writing any file,
#    touching state.json, or sending email. Safe to run repeatedly.
#
#        python weekly_model_pipeline.py --event some-tournament-slug
#    Processes only that one tournament this firing (for real, not test
#    mode) — useful when debugging one event without waiting for the
#    full weekly detection pass or affecting a concurrent event's state.
# ═══════════════════════════════════════════════════════════════════════
